"""
Utilitaires pour la manipulation de fichiers Excel
Lecture, validation, aperçu et analyse
"""

import pandas as pd
import streamlit as st
from pathlib import Path
from typing import List, Optional, Dict, Tuple
import openpyxl
from datetime import datetime


class ExcelFileHandler:
    """Gestionnaire de fichiers Excel pour l'application"""
    
    @staticmethod
    def get_sheet_names(file_path: str) -> List[str]:
        """Retourne la liste des feuilles d'un fichier Excel"""
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True)
            sheet_names = wb.sheetnames
            wb.close()
            return sheet_names
        except Exception as e:
            st.error(f"Erreur lors de la lecture des feuilles : {e}")
            return []
    
    @staticmethod
    def read_excel(file_path: str, sheet_name: Optional[str] = None, nrows: Optional[int] = None) -> Optional[pd.DataFrame]:
        """Lit un fichier Excel et retourne un DataFrame"""
        try:
            if sheet_name:
                df = pd.read_excel(file_path, sheet_name=sheet_name, nrows=nrows)
            else:
                df = pd.read_excel(file_path, nrows=nrows)
            return df
        except Exception as e:
            st.error(f"Erreur lors de la lecture du fichier : {e}")
            return None
    
    @staticmethod
    def get_file_info(file_path: str) -> Dict:
        """Retourne les informations sur un fichier Excel"""
        try:
            path = Path(file_path)
            size = path.stat().st_size
            modified = datetime.fromtimestamp(path.stat().st_mtime)
            
            # Lire le fichier pour obtenir les dimensions
            df = pd.read_excel(file_path, nrows=0)  # Juste les headers
            wb = openpyxl.load_workbook(file_path, read_only=True)
            
            # Compter les lignes de la première feuille
            sheet = wb.active
            num_rows = sheet.max_row
            num_cols = sheet.max_column
            wb.close()
            
            return {
                "name": path.name,
                "size": size,
                "modified": modified,
                "rows": num_rows,
                "columns": num_cols,
                "sheet_names": ExcelFileHandler.get_sheet_names(file_path)
            }
        except Exception as e:
            st.error(f"Erreur lors de la récupération des informations : {e}")
            return {}
    
    @staticmethod
    def preview_dataframe(df: pd.DataFrame, title: str = "Aperçu des données", max_rows: int = 50):
        """Affiche un aperçu d'un DataFrame avec des statistiques"""
        if df is None or df.empty:
            st.warning("Aucune donnée à afficher")
            return
        
        # Informations générales
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Lignes", f"{len(df):,}")
        with col2:
            st.metric("Colonnes", len(df.columns))
        with col3:
            st.metric("Aperçu", f"{min(max_rows, len(df))} lignes")
        
        # Affichage des données
        st.subheader(title)
        st.dataframe(
            df.head(max_rows),
            use_container_width=True,
            height=400
        )
    
    @staticmethod
    def validate_column_exists(df: pd.DataFrame, column_name: str) -> bool:
        """Vérifie si une colonne existe dans le DataFrame"""
        return column_name in df.columns
    
    @staticmethod
    def find_column_matches(df: pd.DataFrame, column_names: List[str]) -> Optional[str]:
        """Trouve la première colonne correspondant à une liste de noms (insensible à la casse)"""
        df_columns_lower = [col.lower() if isinstance(col, str) else str(col).lower() for col in df.columns]
        
        for name in column_names:
            name_lower = name.lower() if isinstance(name, str) else str(name).lower()
            if name_lower in df_columns_lower:
                # Retourner le nom original de la colonne
                index = df_columns_lower.index(name_lower)
                return df.columns[index]
        
        return None
    
    @staticmethod
    def get_column_sample(df: pd.DataFrame, column_name: str, n: int = 5) -> List:
        """Retourne un échantillon de valeurs d'une colonne"""
        if column_name not in df.columns:
            return []
        
        # Retourner les valeurs non-null
        values = df[column_name].dropna().head(n).tolist()
        return values
    
    @staticmethod
    def analyze_matches(source_df: pd.DataFrame, target_df: pd.DataFrame, 
                       source_key: str, target_key: str) -> Dict:
        """Analyse les correspondances possibles entre deux DataFrames"""
        try:
            # Obtenir les clés uniques
            source_keys = set(source_df[source_key].dropna().unique())
            target_keys = set(target_df[target_key].dropna().unique())
            
            # Calculer les correspondances
            matches = source_keys.intersection(target_keys)
            only_in_source = source_keys - target_keys
            only_in_target = target_keys - source_keys
            
            return {
                "total_matches": len(matches),
                "source_total": len(source_keys),
                "target_total": len(target_keys),
                "only_in_source": len(only_in_source),
                "only_in_target": len(only_in_target),
                "match_rate": (len(matches) / len(source_keys) * 100) if len(source_keys) > 0 else 0
            }
        except Exception as e:
            st.error(f"Erreur lors de l'analyse : {e}")
            return {}
    
    @staticmethod
    def save_uploaded_file(uploaded_file) -> Optional[str]:
        """Sauvegarde un fichier uploadé et retourne le chemin absolu"""
        try:
            # Créer un dossier temporaire
            temp_dir = Path("temp_uploads")
            temp_dir.mkdir(parents=True, exist_ok=True)
            
            # Sauvegarder le fichier
            file_path = temp_dir / uploaded_file.name
            with open(file_path, "wb") as f:
                f.write(uploaded_file.getbuffer())
            
            # Retourner le chemin ABSOLU (crucial pour que les classes CellDown/Ticket/SuperXlookup fonctionnent)
            return str(file_path.absolute())
        except Exception as e:
            st.error(f"Erreur lors de la sauvegarde du fichier : {e}")
            return None
    
    @staticmethod
    def format_number(num: int) -> str:
        """Formate un nombre avec des séparateurs de milliers"""
        return f"{num:,}".replace(",", " ")
    
    @staticmethod
    def get_column_names(file_path: str, sheet_name: Optional[str] = None) -> List[str]:
        """Retourne les noms de colonnes d'un fichier Excel"""
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name, nrows=0)
            return df.columns.tolist()
        except Exception as e:
            st.error(f"Erreur lors de la lecture des colonnes : {e}")
            return []
    
    @staticmethod
    def highlight_new_columns(df: pd.DataFrame, original_columns: List[str]) -> pd.DataFrame:
        """Applique un style aux nouvelles colonnes ajoutées"""
        def highlight_cols(col):
            if col.name not in original_columns:
                return ['background-color: #e0e7ff'] * len(col)
            return [''] * len(col)
        
        return df.style.apply(highlight_cols)


def create_date_picker_with_format() -> str:
    """Crée un date picker et retourne la date au format DDMMYYYY"""
    date_input = st.date_input(
        "Date",
        value=datetime.now(),
        format="DD/MM/YYYY"
    )
    
    # Formater en DDMMYYYY
    date_str = date_input.strftime("%d%m%Y")
    st.caption(f"Format généré : `{date_str}`")
    
    return date_str


def display_file_upload_info(uploaded_file):
    """Affiche les informations d'un fichier uploadé"""
    if uploaded_file is not None:
        col1, col2 = st.columns([3, 1])
        with col1:
            st.success(f"✅ Fichier : {uploaded_file.name}")
        with col2:
            size_mb = uploaded_file.size / (1024 * 1024)
            st.caption(f"{size_mb:.2f} MB")
