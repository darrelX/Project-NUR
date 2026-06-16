import pandas as pd
import re
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import Union, List


class SuperXlookup:
    """
    Effectue un XLOOKUP entre deux fichiers Excel (source et cible, une seule feuille chacun).
    Ajoute un préfixe "[reference_name - date_aujourd'hui]" devant chaque valeur trouvée.
    
    Paramètres :
        source_file_path     : chemin du fichier source (sera modifié)
        target_file_path     : chemin du fichier cible (lu seulement)
        source_key_column    : colonne clé dans la source (lettre, numéro ou nom)
        target_key_column    : colonne clé dans la cible
        target_value_column  : colonne dont on veut rapporter la valeur
        result_position_column : colonne où écrire le résultat (ex: 'C', '5', 'last_column')
        result_column_name   : nom (en-tête) de la colonne résultat
        source_sheet_name    : nom de la feuille source (défaut = première feuille)
        target_sheet_name    : nom de la feuille cible (défaut = première feuille)
        reference_name       : nom de référence pour le préfixe (ex: "NOKIA")
    """
    def __init__(
        self,
        source_file_path: str,
        target_file_path: str,
        source_key_column: Union[str, List[str]],  # Nom(s) de colonne ou lettre
        target_key_column: Union[str, List[str]],  # Nom(s) de colonne ou lettre
        target_value_column: Union[str, List[str]],  # Nom(s) de colonne ou lettre
        result_position_column: str,
        result_column_name: str,
        source_sheet_name: str = "",
        target_sheet_name: str = "",
        reference_name: str = "",
        reference_date: str = "",   # <-- date de référence (format : "DD/MM/YYYY")
    ):
        self.source_file_path = source_file_path
        self.target_file_path = target_file_path
        self.source_key_column = source_key_column
        self.target_key_column = target_key_column
        self.target_value_column = target_value_column
        self.result_position_column = result_position_column
        self.result_column_name = result_column_name.strip()
        self.source_sheet_name = source_sheet_name if source_sheet_name else 0
        self.target_sheet_name = target_sheet_name if target_sheet_name else 0
        self.reference_name = reference_name.strip()
        self.reference_date = reference_date.strip() if reference_date else datetime.now().strftime("%d/%m/%Y")

        # Vérifications d'existence
        if not Path(source_file_path).exists():
            raise FileNotFoundError(f"Fichier source introuvable : {source_file_path}")
        if not Path(target_file_path).exists():
            raise FileNotFoundError(f"Fichier cible introuvable : {target_file_path}")

    # ------------------------------------------------------------------
    # Utilitaires de conversion lettres Excel ↔ indices
    # ------------------------------------------------------------------
    @staticmethod
    def _col_letter_to_index(col: str) -> int:
        """Convertit une lettre Excel (A, B, AA) en index 0-based."""
        col = col.upper().strip()
        result = 0
        for ch in col:
            result = result * 26 + (ord(ch) - ord('A') + 1)
        return result - 1

    @staticmethod
    def _index_to_col_letter(idx: int) -> str:
        """Convertit un index 0-based en lettre Excel."""
        result = ""
        idx += 1
        while idx > 0:
            idx, remainder = divmod(idx - 1, 26)
            result = chr(65 + remainder) + result
        return result

    def _find_column_by_names(self, df: pd.DataFrame, column_names: Union[str, List[str]]) -> str:
        """
        Trouve une colonne par nom(s), insensible à la casse.
        Si une liste est fournie, retourne la première colonne trouvée.
        """
        # Convertir en liste si c'est une chaîne
        if isinstance(column_names, str):
            column_names = [column_names]
        
        # Créer un mapping nom_lowercase -> nom_réel
        col_map = {str(col).strip().lower(): col for col in df.columns}
        
        # Chercher la première correspondance
        for name in column_names:
            name_lower = str(name).strip().lower()
            if name_lower in col_map:
                return col_map[name_lower]
        
        # Aucune correspondance trouvée
        raise KeyError(f"Aucune des colonnes {column_names} n'a été trouvée")
    
    def _resolve_column_name(self, df: pd.DataFrame, column_spec: Union[str, List[str]]) -> str:
        """
        Retourne le nom réel de la colonne (l'en-tête) à partir d'une spécification.
        La spécification peut être :
          - une liste de noms : ['Site ID', 'SITE_ID', 'site id'] (insensible à la casse)
          - un nom de colonne : 'Code', 'Nom' (insensible à la casse)
          - une lettre Excel : 'A', 'B', 'AA'
          - un numéro (1-based) : '1', '2', '3'
        """
        # Si c'est une liste, chercher par noms
        if isinstance(column_spec, list):
            return self._find_column_by_names(df, column_spec)
        
        spec = str(column_spec).strip()

        # 1. Vérifier si c'est un index numérique
        if spec.isdigit():
            idx = int(spec) - 1
            if 0 <= idx < len(df.columns):
                return df.columns[idx]
            raise IndexError(f"Colonne numéro {spec} hors limites (max {len(df.columns)})")

        # 2. Chercher par nom (insensible à la casse)
        try:
            return self._find_column_by_names(df, spec)
        except KeyError:
            pass

        # 3. Vérifier si c'est une lettre de colonne (A, B, AA...)
        # Limiter à 1-2 caractères pour éviter confusion avec noms
        if spec.isalpha() and len(spec) <= 2:
            idx = self._col_letter_to_index(spec)
            if 0 <= idx < len(df.columns):
                return df.columns[idx]
            raise IndexError(f"Colonne lettre {spec} hors limites")

        # 4. Si rien ne correspond
        raise KeyError(f"Colonne '{spec}' introuvable")

    def _determine_result_column_position(self, df_source: pd.DataFrame) -> int:
        """
        Retourne l'index (0-based) de la colonne où écrire le résultat, en fonction
        de result_position_column.
        Si la valeur est 'last_column' (ou variante), l'index est len(df_source.columns)
        (i.e. nouvelle colonne à droite).
        Sinon, on convertit la spécification (lettre ou nombre) en index.
        """
        pos = str(self.result_position_column).strip().lower()
        if pos in ("last_column", "derniere_libre", "dernière_libre"):
            return len(df_source.columns)   # nouvelle colonne après la dernière
        # Position normale : lettre ou nombre
        if pos.isdigit():
            return int(pos) - 1
        if pos.isalpha():
            return self._col_letter_to_index(pos)
        # Cas particulier : si c'est déjà un nom d'en-tête, on cherche sa position
        if pos in df_source.columns:
            return df_source.columns.get_loc(pos)
        raise ValueError(f"Position de colonne invalide : '{self.result_position_column}'")

    def _save_with_formatting(self, df_source: pd.DataFrame) -> None:
        """
        Ajoute uniquement les NOUVELLES colonnes à droite du fichier Excel.
        Conserve TOUTES les données existantes sans y toucher.
        """
        import numpy as np
        
        # Charger le workbook existant
        wb = load_workbook(self.source_file_path)
        
        # Déterminer la feuille à modifier
        if isinstance(self.source_sheet_name, str) and self.source_sheet_name:
            ws = wb[self.source_sheet_name]
        else:
            ws = wb.active
        
        # Identifier les colonnes existantes dans Excel
        existing_cols = []
        for col_idx in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col_idx).value
            if header:
                existing_cols.append(header)
        
        # Identifier les NOUVELLES colonnes (celles à ajouter)
        new_columns = []
        new_col_indices = []
        for idx, col_name in enumerate(df_source.columns):
            if col_name not in existing_cols:
                new_columns.append(col_name)
                new_col_indices.append(idx)
        
        if not new_columns:
            print("✓ Aucune nouvelle colonne à ajouter")
            wb.close()
            return
        
        # Ajouter les en-têtes des nouvelles colonnes à droite
        start_col = len(existing_cols) + 1
        for i, col_name in enumerate(new_columns):
            ws.cell(row=1, column=start_col + i, value=col_name)
        
        # Écrire UNIQUEMENT les données des nouvelles colonnes
        for row_idx, row_data in enumerate(df_source.values, start=2):
            for i, df_col_idx in enumerate(new_col_indices):
                value = row_data[df_col_idx]
                
                # Convertir les types pandas/numpy en types Python natifs
                if pd.isna(value):
                    cell_value = None
                elif isinstance(value, (np.integer, np.int64, np.int32)):
                    cell_value = int(value)
                elif isinstance(value, (np.floating, np.float64, np.float32)):
                    cell_value = float(value)
                elif isinstance(value, np.bool_):
                    cell_value = bool(value)
                else:
                    cell_value = value
                
                ws.cell(row=row_idx, column=start_col + i, value=cell_value)
        
        # Sauvegarder et fermer
        wb.save(self.source_file_path)
        wb.close()
        print(f"🎨 {len(new_columns)} nouvelle(s) colonne(s) ajoutée(s) : {', '.join(new_columns)}")

    def run(self) -> None:
        """Exécute le XLOOKUP et sauvegarde le fichier source."""
        print("🔍 Démarrage XLOOKUP...")

        # Chargement
        df_source = pd.read_excel(self.source_file_path, sheet_name=self.source_sheet_name)
        df_target = pd.read_excel(self.target_file_path, sheet_name=self.target_sheet_name)

        # Résolution des colonnes clés et valeur
        source_key_col = self._resolve_column_name(df_source, self.source_key_column)
        target_key_col = self._resolve_column_name(df_target, self.target_key_column)
        target_value_col = self._resolve_column_name(df_target, self.target_value_column)

        # Détermination de la colonne résultat - utiliser le nom du fichier cible
        target_filename = Path(self.target_file_path).name  # Nom complet du fichier avec extension
        result_col_name = target_filename
        target_idx = self._determine_result_column_position(df_source)

        # Supprimer toutes les colonnes existantes avec ce nom (éviter les doublons)
        if result_col_name in df_source.columns:
            df_source = df_source.loc[:, ~df_source.columns.duplicated()]
            cols_to_drop = [col for col in df_source.columns if col == result_col_name]
            df_source = df_source.drop(columns=cols_to_drop)
            # Recalculer target_idx après suppression
            target_idx = self._determine_result_column_position(df_source)

        # Si l'index est >= nombre de colonnes, on ajoute des colonnes vides jusqu'à cet index
        if target_idx >= len(df_source.columns):
            nb_new_cols = target_idx - len(df_source.columns) + 1
            for _ in range(nb_new_cols):
                df_source[f"tmp_{_}"] = ""   # colonnes provisoires
            # Renommer la dernière colonne ajoutée avec le nom voulu
            df_source.rename(columns={df_source.columns[target_idx]: result_col_name}, inplace=True)
        else:
            # La colonne cible existe déjà (par son index) : on l'utilise
            existing_name = df_source.columns[target_idx]
            if existing_name != result_col_name:
                # Renommer la colonne existante
                df_source.rename(columns={existing_name: result_col_name}, inplace=True)

        # Nettoyage des clés (suppression des espaces, conversion en chaîne)
        df_source[source_key_col] = df_source[source_key_col].astype(str).str.strip()
        df_target[target_key_col] = df_target[target_key_col].astype(str).str.strip()

        # Mapping
        mapping = dict(zip(df_target[target_key_col], df_target[target_value_col]))
        df_source[result_col_name] = df_source[source_key_col].map(mapping)

        # Remplacer les NaN par chaîne vide (plus propre qu'un espace)
        df_source[result_col_name].fillna("", inplace=True)

        # --- Ajout du préfixe [reference_name - date_extraite_du_fichier] ---
        if self.reference_name:
            # Extraire le nom du fichier cible (sans le chemin)
            target_filename = Path(self.target_file_path).name
            
            # Chercher un pattern de date dans le nom du fichier
            # Pattern 1: avec slashes (jj/mm/aaaa ou j/m/aaaa) + heure optionnelle
            date_pattern_slash = r'(\d{1,2}/\d{1,2}/\d{4}(?:\s+\d{1,2}[hH](?:\d{2})?)?(?:\s+[A-Z]+)?)'
            # Pattern 2: avec tirets (jj-mm-aaaa ou j-m-aaaa) + heure optionnelle
            date_pattern_dash = r'(\d{1,2}-\d{1,2}-\d{4}(?:\s+\d{1,2}[hH](?:\d{2})?)?(?:\s+[A-Z]+)?)'
            # Pattern 3: sans séparateurs (jjmmaaaa) + heure optionnelle
            date_pattern_no_sep = r'(\d{8}(?:\s+\d{1,2}[hH](?:\d{2})?)?(?:\s+[A-Z]+)?)'
            
            match = re.search(date_pattern_slash, target_filename)
            if not match:
                match = re.search(date_pattern_dash, target_filename)
            if not match:
                match = re.search(date_pattern_no_sep, target_filename)
            
            if match:
                file_date_info = match.group(1).strip()
                prefix = f"{{{self.reference_name} {file_date_info}}} : "
            else:
                # Si pas de date trouvée, utiliser la date de référence par défaut
                prefix = f"{{{self.reference_name} - {self.reference_date}}} : "
            
            # Appliquer le préfixe uniquement aux cellules non vides
            df_source[result_col_name] = df_source[result_col_name].apply(
                lambda x: f"{prefix}{x}" if x != "" else ""
            )

        # Statistiques (nombre de cellules non vides après ajout du préfixe)
        matched = (df_source[result_col_name] != "").sum()
        total = len(df_source)
        print(f"✅ Correspondances : {matched}/{total}")

        # Sauvegarde avec conservation de la mise en forme
        self._save_with_formatting(df_source)
        print(f"💾 Fichier source mis à jour : {self.source_file_path}")


# =============================================================================
# TESTS
# =============================================================================
if __name__ == "__main__":


    # Test avec "last_column" et sans reference_name
    print("\n" + "=" * 60)
    print("TEST XLOOKUP (dernière colonne libre, sans préfixe)")
    print("=" * 60)
    xl2 = SuperXlookup(
        source_file_path=r"C:\Users\f50056342\Desktop\computer science\NUR Project Lyne\inputs\Book1.xlsx",
        target_file_path=r"C:\Users\f50056342\Desktop\my work\OCM RAN\OCM RAN INCIDENT FOLOW-UP 23-05-2026 00H UTC.xlsx",
        source_key_column="B",
        target_key_column="D",
        target_value_column="J",
        result_position_column="last_column",
        result_column_name="OCM RAN 23-05-2026 00H UTC",
        source_sheet_name="Sheet1",
        target_sheet_name="ALL SITES DOWN",
        reference_name=""                # pas de préfixe
    )
    xl2.run()

    # Affichage du résultat final
    df = pd.read_excel(r"C:\Users\f50056342\Desktop\computer science\NUR Project Lyne\inputs\Book1.xlsx")
    print("\n📄 Aperçu du fichier source après exécution :")
    print(df.head(30))