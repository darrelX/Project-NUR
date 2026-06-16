import os
import re
from pathlib import Path
from typing import List, Optional, Dict, Any, Union
from dataclasses import dataclass, field
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


@dataclass
class CellDown:
    """
    Classe pour manipuler des fichiers Excel : affichage des feuilles,
    filtrage par date, et recherche XLOOKUP-like.
    """

    source_file_path: str
    target_file_path: str
    colown_key_path_source: Union[str, List[str]]  # colonne clé dans la source (nom(s) ou lettre)
    target_value_column: Union[str, List[str]]     # colonne à rapporter depuis la cible
    result_position_column: str   # colonne à remplir dans la source
    source_sheet_path: str = ""   # nom de la feuille source
    selected_date: str = ""       # date éventuellement filtrée
    reference_name: str = ""      # nom de référence pour les résultats (auto-généré si vide)
    reference_date: str = ""      # date de référence (format : "DD/MM/YYYY")

    # Nouveaux attributs pour la recherche croisée
    target_key_column: Optional[Union[str, List[str]]] = None   # colonne clé dans la cible
    target_sheet_name: Optional[str] = None

    # Paramètres de date et colonne de départ
    date_str: str = ""
    start_column: str = "last_column"

    _source_path: Path = field(init=False, repr=False)
    _target_path: Path = field(init=False, repr=False)

    def __post_init__(self):
        self._source_path = Path(self.source_file_path)
        self._target_path = Path(self.target_file_path)
        if self.target_key_column is None:
            self.target_key_column = self.colown_key_path_source
        # Date de référence par défaut = date du jour
        if not self.reference_date:
            self.reference_date = datetime.now().strftime("%d/%m/%Y")
        # Génération automatique du reference_name à partir du fichier cible
        if not self.reference_name:
            self.reference_name = self._extract_reference_from_filename(self.target_file_path)
            if self.reference_name:
                print(f"🔍 reference_name auto-généré : '{self.reference_name}'")
            else:
                print("⚠️ Impossible d'extraire reference_name du nom du fichier, préfixe désactivé.")

    # ----------------------------------------------------------------------
    # Extraction automatique du nom de référence
    # ----------------------------------------------------------------------
    def _extract_reference_from_filename(self, file_path: str) -> str:
        """
        Extrait un nom de référence (ex: 'CellDown Nokia') à partir du nom du fichier cible.
        Logique : récupère le nom de base sans extension, cherche un motif connu.
        """
        base_name = Path(file_path).stem  # sans chemin ni extension
        
        # Patterns possibles (par ordre de priorité)
        patterns = [
            # 1. Après 'CELLS_DOWN_' ou 'CELL_DOWN_' (insensible à la casse)
            r'(?i)(?:CELLS?[_ ]DOWN[_ ]+)([A-Za-z0-9]+)',
            # 2. Avant le premier nombre à 8 chiffres (date jjmmaaaa)
            r'^([A-Za-z_]+)(?=\d{8})',
            # 3. Dernier mot alphabétique avant une date
            r'([A-Za-z]+)(?=\s*\d{8})',
            # 4. Tout mot en majuscules de 3 lettres ou plus (ex: NOKIA, HUAWEI)
            r'\b([A-Z]{3,})\b'
        ]
        
        for pattern in patterns:
            match = re.search(pattern, base_name, re.IGNORECASE)
            if match:
                vendor = match.group(1).strip().title()
                return f"CellDown {vendor}"
        
        # Si rien ne correspond, on prend le premier mot significatif
        words = re.findall(r'[A-Za-z]+', base_name)
        if words:
            # Éviter les mots trop courts ou génériques
            for w in words:
                if len(w) >= 3 and w.upper() not in ('DAILY', 'WEEK', 'CELL', 'DOWN', 'CELLS', 'FILE'):
                    return f"CellDown {w.title()}"
        
        # Dernier recours : utiliser le nom complet du fichier (sans extension)
        return f"CellDown {base_name.replace('_', ' ').title()}"

    # ----------------------------------------------------------------------
    # Utilitaires
    # ----------------------------------------------------------------------
    def _col_letter_to_index(self, col: str) -> int:
        col = col.upper().strip()
        result = 0
        for char in col:
            result = result * 26 + (ord(char) - ord('A') + 1)
        return result - 1

    def _find_column_by_names(self, df: pd.DataFrame, column_names: Union[str, List[str]]) -> str:
        """
        Trouve une colonne par nom(s), insensible à la casse.
        Si une liste est fournie, retourne la première colonne trouvée.
        """
        # Convertir en liste si c'est une chaîne
        if isinstance(column_names, str):
            column_names = [column_names]
        
        # Créer un mapping nom_lowercase -> nom_réel
        col_map = {str(col).strip().lower(): col for col in df.columns}
        
        # Chercher la première correspondance
        for name in column_names:
            name_lower = str(name).strip().lower()
            if name_lower in col_map:
                return col_map[name_lower]
        
        # Aucune correspondance trouvée
        raise KeyError(f"Aucune des colonnes {column_names} n'a été trouvée")
    
    def _get_column_by_position(self, df: pd.DataFrame, position: Union[str, List[str]]) -> str:
        # Si c'est une liste, chercher par noms
        if isinstance(position, list):
            return self._find_column_by_names(df, position)
        
        position = str(position).strip()
        
        # 1. Vérifier si c'est un index numérique (1, 2, 3...)
        if position.isdigit():
            idx = int(position) - 1
            if 0 <= idx < len(df.columns):
                return df.columns[idx]
            raise IndexError(f"Position {position} hors limites")
        
        # 2. Chercher par nom (insensible à la casse)
        try:
            return self._find_column_by_names(df, position)
        except KeyError:
            pass
        
        # 3. Vérifier si c'est une lettre de colonne (A, B, C...)
        # Uniquement si 1-2 caractères pour éviter de confondre avec des noms
        if position.isalpha() and len(position) <= 2:
            idx = self._col_letter_to_index(position)
            if 0 <= idx < len(df.columns):
                return df.columns[idx]
            raise IndexError(f"Colonne {position} hors limites")
        
        # 4. Si rien ne correspond
        raise KeyError(f"Colonne '{position}' introuvable")

    def _index_to_col_letter(self, idx: int) -> str:
        result = ""
        idx += 1
        while idx > 0:
            idx, remainder = divmod(idx - 1, 26)
            result = chr(65 + remainder) + result
        return result

    def _read_sheet_safe(self, file_path: str, sheet_name: str) -> pd.DataFrame:
        """
        Lit une feuille Excel de manière robuste face aux filtres invalides.
        On force dtype=str pour éviter les erreurs de validation des filtres.
        """
        return pd.read_excel(
            file_path,
            sheet_name=sheet_name,
            engine='openpyxl',
            dtype=str,
            keep_default_na=False,
            na_values=['']
        )

    def _save_with_formatting(self, df_source: pd.DataFrame) -> None:
        """
        Ajoute uniquement les NOUVELLES colonnes à droite du fichier Excel.
        Conserve TOUTES les données existantes sans y toucher.
        """
        import numpy as np
        
        # Charger le workbook existant
        wb = load_workbook(self.source_file_path)
        
        # Déterminer la feuille à modifier
        if isinstance(self.source_sheet_path, str) and self.source_sheet_path:
            ws = wb[self.source_sheet_path]
        else:
            ws = wb.active
        
        # Identifier les colonnes existantes dans Excel
        existing_cols = []
        for col_idx in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col_idx).value
            if header:
                existing_cols.append(header)
        
        # Identifier les NOUVELLES colonnes (celles à ajouter)
        new_columns = []
        new_col_indices = []
        for idx, col_name in enumerate(df_source.columns):
            if col_name not in existing_cols:
                new_columns.append(col_name)
                new_col_indices.append(idx)
        
        if not new_columns:
            print("✓ Aucune nouvelle colonne à ajouter")
            wb.close()
            return
        
        # Ajouter les en-têtes des nouvelles colonnes à droite
        start_col = len(existing_cols) + 1
        for i, col_name in enumerate(new_columns):
            ws.cell(row=1, column=start_col + i, value=col_name)
        
        # Écrire UNIQUEMENT les données des nouvelles colonnes
        for row_idx, row_data in enumerate(df_source.values, start=2):
            for i, df_col_idx in enumerate(new_col_indices):
                value = row_data[df_col_idx]
                
                # Convertir les types pandas/numpy en types Python natifs
                if pd.isna(value):
                    cell_value = None
                elif isinstance(value, (np.integer, np.int64, np.int32)):
                    cell_value = int(value)
                elif isinstance(value, (np.floating, np.float64, np.float32)):
                    cell_value = float(value)
                elif isinstance(value, np.bool_):
                    cell_value = bool(value)
                else:
                    cell_value = value
                
                ws.cell(row=row_idx, column=start_col + i, value=cell_value)
        
        # Sauvegarder et fermer
        wb.save(self.source_file_path)
        wb.close()
        print(f"🎨 {len(new_columns)} nouvelle(s) colonne(s) ajoutée(s) : {', '.join(new_columns)}")

    # ----------------------------------------------------------------------
    # Méthode principale optimisée
    # ----------------------------------------------------------------------
    def super_xlookup_par_date(self) -> None:
        """
        Filtre les feuilles par date et effectue un XLOOKUP pour chaque feuille.
        Lit uniquement les feuilles nécessaires (pas de chargement complet du fichier).
        Ajoute le préfixe "[reference_name - date_aujourd'hui] " devant chaque texte.
        """
        if not self.date_str:
            raise ValueError("L'attribut 'date_str' doit être renseigné.")

        # --- Récupérer les noms des feuilles (rapide, pas de données) ---
        try:
            excel_file = pd.ExcelFile(self.target_file_path, engine='openpyxl')
            all_sheets = excel_file.sheet_names
        except Exception as e:
            print(f"❌ Impossible de lire les feuilles du fichier cible : {e}")
            return

        # --- Filtrer par date ---
        try:
            date_obj = datetime.strptime(self.date_str, "%d%m%Y")
            pattern = date_obj.strftime("%d%m%Y")
        except ValueError:
            print("Format de date invalide. Utilisez jjmmaaaa (ex: 12032025).")
            return

        feuilles_filtrees = [s for s in all_sheets if pattern in str(s)]

        if not feuilles_filtrees:
            print(f"❌ Aucune feuille ne contient la date '{pattern}' dans {self.target_file_path}")
            return

        print(f"📅 Feuilles trouvées pour la date '{pattern}':")
        for i, feuille in enumerate(feuilles_filtrees, 1):
            print(f"   {i}. {feuille}")

        # --- Charger le fichier source (une seule fois) ---
        df_source = self._read_sheet_safe(self.source_file_path, sheet_name=self.source_sheet_path or 0)
        
        # --- Déterminer la colonne de départ ---
        start_col = str(self.start_column).strip().lower()
        if start_col in ("last_column", "derniere_libre", "dernière_libre"):
            # Commencer après la dernière colonne existante
            start_idx = len(df_source.columns)
        elif start_col.isdigit():
            start_idx = int(start_col) - 1
        elif start_col.isalpha():
            start_idx = self._col_letter_to_index(start_col.upper())
        else:
            start_idx = 2  # par défaut colonne C
        source_key_col = self._get_column_by_position(df_source, self.colown_key_path_source)
        df_source[source_key_col] = df_source[source_key_col].astype(str).str.strip()

        # Supprimer les colonnes existantes avec les noms des feuilles filtrées (éviter les doublons)
        df_source = df_source.loc[:, ~df_source.columns.duplicated()]
        cols_to_drop = [col for col in df_source.columns if col in feuilles_filtrees]
        if cols_to_drop:
            df_source = df_source.drop(columns=cols_to_drop)
            print(f"🗑️  Colonnes existantes supprimées : {cols_to_drop}")

        # --- Parcourir chaque feuille filtrée (lecture à la demande) ---
        for i, feuille in enumerate(feuilles_filtrees):
            current_col_idx = start_idx + i
            current_col_letter = self._index_to_col_letter(current_col_idx)

            print(f"   [{i+1}/{len(feuilles_filtrees)}] Feuille '{feuille}' → Colonne {current_col_letter}")
            
            # Extraire la date/heure du nom de la feuille pour le préfixe
            sheet_date_info = ""
            if self.reference_name:
                # Chercher un pattern de date dans le nom de la feuille
                # Pattern 1: avec slashes (jj/mm/aaaa ou j/m/aaaa) + heure optionnelle
                date_pattern_slash = r'(\d{1,2}/\d{1,2}/\d{4}(?:\s+\d{1,2}h(?:\d{2})?)?)'
                # Pattern 2: sans slashes (jjmmaaaa) + heure optionnelle
                date_pattern_no_slash = r'(\d{8}(?:\s+\d{1,2}h(?:\d{2})?)?)'
                
                match = re.search(date_pattern_slash, feuille)
                if not match:
                    match = re.search(date_pattern_no_slash, feuille)
                
                if match:
                    sheet_date_info = match.group(1)
                    prefix = f"{{{self.reference_name} {sheet_date_info}}} : "
                else:
                    # Si pas de date trouvée, utiliser la date de référence par défaut
                    prefix = f"{{{self.reference_name} - {self.reference_date}}} : "
            else:
                prefix = ""

            try:
                df_target = self._read_sheet_safe(self.target_file_path, sheet_name=feuille)
            except Exception as e:
                print(f"      ❌ Erreur de lecture de la feuille : {e}")
                df_source[feuille] = "ERREUR_LECTURE"
                continue

            if df_target.empty:
                print("      ⚠️ Feuille vide, colonne remplie avec chaînes vides.")
                df_source[feuille] = ""
                continue

            # Résolution des colonnes cibles
            try:
                target_key_input = self.target_key_column or self.colown_key_path_source
                target_key_col = self._get_column_by_position(df_target, target_key_input)
                target_value_col = self._get_column_by_position(df_target, self.target_value_column)
            except (IndexError, KeyError) as e:
                print(f"      ⚠️ Colonne manquante dans cette feuille ({e}), ignorée.")
                df_source[feuille] = ""
                continue

            df_target[target_key_col] = df_target[target_key_col].astype(str).str.strip()
            mapping = dict(zip(df_target[target_key_col], df_target[target_value_col]))

            # XLOOKUP
            df_source[feuille] = df_source[source_key_col].map(mapping).fillna("")

            # Ajout du préfixe si demandé
            if prefix:
                df_source[feuille] = df_source[feuille].apply(lambda x: f"{prefix}{x}" if x != "" else "")

            matched = (df_source[feuille] != "").sum()
            print(f"      Correspondances (non vides) : {matched}/{len(df_source)}")

        # --- Sauvegarde avec conservation de la mise en forme ---
        self._save_with_formatting(df_source)
        print(f"\n✅ Mise à jour terminée : {self.source_file_path}")
        print(f"   Colonnes ajoutées : {self._index_to_col_letter(start_idx)} à {self._index_to_col_letter(start_idx + len(feuilles_filtrees) - 1)}")


if __name__ == "__main__":
    # Exemple d'utilisation : reference_name est auto-généré à partir du fichier cible
    celldown = CellDown(
        source_file_path=r"C:\Users\f50056342\Desktop\computer science\NUR Project Lyne\inputs\Book1.xlsx",
        target_file_path=r"C:\Users\f50056342\Desktop\my work\cellsdow files\celldown Nokia\DAILY_W24_CELLS_DOWN_NOKIA_0806 AU 11062026 18h10,.xlsx",
        colown_key_path_source="Codesite",
        target_key_column=["Site Code", "SITE_CODE", "site code", "Code Site"],
        target_value_column=["Comment", "COMMENT", "comment"],
        result_position_column="last_column",
        source_sheet_path="Sheet1",
        date_str="11062026",
    )
    celldown.super_xlookup_par_date()

    # Affichage du résultat
    df = pd.read_excel(r"C:\Users\f50056342\Desktop\computer science\NUR Project Lyne\inputs\Book1.xlsx")
    print("\nAperçu du fichier résultat :")
    print(df.head(30))