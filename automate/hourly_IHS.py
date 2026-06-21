import os
from pathlib import Path
from typing import List, Optional, Any, Union
from dataclasses import dataclass, field
import pandas as pd
from datetime import datetime
import re
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


@dataclass
class HourlyIHS:
    """
    Classe pour manipuler des fichiers Excel :
    - Extraction automatique de codes sites (préfixes définis) dans une colonne cible.
    - XLOOKUP entre source et cible.
    - Possibilité de concaténer plusieurs colonnes cibles (TEXTJOIN) au lieu d'une seule.
    - Ajout d'un préfixe [reference_name - date] devant les résultats.
    - Suppression automatique des filtres Excel pour éviter les bugs.

    Les paramètres de colonnes peuvent être :
        - une lettre Excel (ex: "A", "B", "AA")
        - un numéro de colonne (1-indexé)
        - le nom exact de la colonne (insensible à la casse)
        - une liste de noms possibles (pour la recherche)
    """

    source_file_path: str
    target_file_path: str
    source_key_column: Union[str, List[str]]  # colonne clé dans la source
    result_position_column: str               # colonne où écrire le résultat
    result_column_name: str                  # nom de la colonne résultat
    source_sheet_name: str = ""              # feuille source (défaut = première)
    target_sheet_name: str = ""              # feuille cible (défaut = première)
    reference_name: str = ""                 # pour le préfixe [NOM - date]
    reference_date: Optional[str] = None     # date pour le préfixe (si None, date du jour)

    # Paramètres pour la valeur à rapporter depuis la cible
    target_value_column: Optional[Union[str, List[str]]] = None   # colonne unique
    target_join_columns: Optional[List[Union[str, List[str]]]] = None  # liste des colonnes à concaténer
    join_separator: str = ".."                # séparateur pour TEXTJOIN
    ignore_empty: bool = True                 # ignorer les cellules vides

    # Paramètres pour l'extraction automatique de la clé cible
    target_key_column: Optional[Union[str, List[str]]] = None   # colonne clé cible
    extract_source_column: Optional[Union[str, List[str]]] = None  # colonne source pour extraire les codes
    extract_prefixes: List[str] = field(default_factory=lambda: ["ABC"])  # préfixes à rechercher (ex: "ABC", "DEF")

    # ------------------------------------------------------------------
    # Suppression des filtres Excel
    # ------------------------------------------------------------------
    @staticmethod
    def _remove_auto_filters(file_path: str, sheet_name: Optional[str] = None) -> None:
        try:
            wb = load_workbook(file_path)
            if sheet_name is None:
                target_sheet = wb.worksheets[0]
            else:
                target_sheet = wb[sheet_name]
            if target_sheet.auto_filter:
                target_sheet.auto_filter = None
                wb.save(file_path)
                print(f"🧹 Filtre supprimé sur la feuille '{target_sheet.title}' de {file_path}")
        except Exception as e:
            print(f"⚠️ Impossible de supprimer le filtre de {file_path} (feuille {sheet_name}) : {e}")

    # ------------------------------------------------------------------
    # Utilitaires de conversion lettres Excel ↔ indices
    # ------------------------------------------------------------------
    @staticmethod
    def _col_letter_to_index(col: str) -> int:
        col = col.upper().strip()
        result = 0
        for ch in col:
            result = result * 26 + (ord(ch) - ord('A') + 1)
        return result - 1

    @staticmethod
    def _index_to_col_letter(idx: int) -> str:
        result = ""
        idx += 1
        while idx > 0:
            idx, remainder = divmod(idx - 1, 26)
            result = chr(65 + remainder) + result
        return result

    # ---------- RECHERCHE AMÉLIORÉE DE COLONNES ----------
    def _find_column_by_names(self, df: pd.DataFrame, column_names: Union[str, List[str]]) -> str:
        """
        Recherche une colonne parmi une liste de noms possibles.
        - Normalisation : minuscules, suppression des espaces.
        - D'abord correspondance exacte, puis correspondance partielle.
        """
        if isinstance(column_names, str):
            column_names = [column_names]

        # Normaliser les noms de colonnes du DataFrame : minuscule, sans espaces
        col_map = {}
        for col in df.columns:
            col_norm = str(col).strip().lower().replace(" ", "")
            col_map[col_norm] = col

        # 1. Correspondance exacte (normalisée)
        for name in column_names:
            name_norm = str(name).strip().lower().replace(" ", "")
            if name_norm in col_map:
                return col_map[name_norm]

        # 2. Correspondance partielle (le terme cherché est contenu dans un nom de colonne, ou inversement)
        for name in column_names:
            name_norm = str(name).strip().lower().replace(" ", "")
            for col_norm, col_orig in col_map.items():
                if name_norm in col_norm or col_norm in name_norm:
                    return col_orig

        raise KeyError(f"Aucune des colonnes {column_names} n'a été trouvée parmi {list(df.columns)}")
    
    def _resolve_column_name(self, df: pd.DataFrame, column_spec: Union[str, List[str]]) -> str:
        if isinstance(column_spec, list):
            return self._find_column_by_names(df, column_spec)
        spec = str(column_spec).strip()
        if spec.isdigit():
            idx = int(spec) - 1
            if 0 <= idx < len(df.columns):
                return df.columns[idx]
            raise IndexError(f"Colonne numéro {spec} hors limites")
        try:
            return self._find_column_by_names(df, spec)
        except KeyError:
            pass
        if spec.isalpha() and len(spec) <= 2:
            idx = self._col_letter_to_index(spec)
            if 0 <= idx < len(df.columns):
                return df.columns[idx]
            raise IndexError(f"Colonne lettre {spec} hors limites")
        raise KeyError(f"Colonne '{spec}' introuvable")

    def _resolve_columns_list(self, df: pd.DataFrame, columns_spec: List[Union[str, List[str]]]) -> List[str]:
        result = []
        for col_spec in columns_spec:
            try:
                col_name = self._resolve_column_name(df, col_spec)
                result.append(col_name)
            except (KeyError, IndexError):
                print(f"⚠️ Colonne spécifiée '{col_spec}' non trouvée dans le fichier cible, ignorée.")
                continue
        if not result:
            print("⚠️ Aucune colonne valide trouvée parmi target_join_columns, le résultat sera vide.")
        return result

    def _determine_result_column_position(self, df_source: pd.DataFrame) -> int:
        pos = str(self.result_position_column).strip().lower()
        # 1. Mots-clés spéciaux
        if pos in ("last_free", "derniere_libre", "dernière_libre", "last_column", "derniere_colonne", "dernière_colonne"):
            return len(df_source.columns)
        # 2. Nom de colonne existant (priorité sur lettre)
        if pos in df_source.columns:
            return df_source.columns.get_loc(pos)
        # 3. Numéro de colonne
        if pos.isdigit():
            return int(pos) - 1
        # 4. Lettre Excel (seulement si la chaîne est alphabétique)
        if pos.isalpha():
            return self._col_letter_to_index(pos)
        raise ValueError(f"Position de colonne invalide : '{self.result_position_column}'")

    # ================================================================
    # Sauvegarde avec préservation des styles et de l'ordre
    # ================================================================
    def _save_with_formatting(self, df_source: pd.DataFrame) -> None:
        """
        Sauvegarde le DataFrame source dans le fichier Excel en préservant :
        - l'ordre des colonnes (selon df_source)
        - les positions des colonnes existantes (réutilisation)
        - les styles des cellules existantes (non écrasés)
        - les valeurs mises à jour pour toutes les colonnes
        """
        import numpy as np

        wb = load_workbook(self.source_file_path)
        if self.source_sheet_name:
            ws = wb[self.source_sheet_name]
        else:
            ws = wb.active

        # 1. Récupérer les en-têtes existants et leur position (1-indexé)
        existing_headers = {}
        for col_idx in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col_idx).value
            if header is not None:
                existing_headers[header] = col_idx

        # 2. Déterminer la prochaine colonne libre (pour les nouvelles colonnes)
        next_col = ws.max_column + 1

        # 3. Parcourir les colonnes de df_source dans l'ordre
        for col_idx, col_name in enumerate(df_source.columns):
            # Trouver la position dans le workbook
            if col_name in existing_headers:
                wb_col = existing_headers[col_name]
            else:
                wb_col = next_col
                next_col += 1
                # Écrire l'en-tête
                ws.cell(row=1, column=wb_col, value=col_name)

            # Écrire les données pour cette colonne (ligne 2 à n)
            for i, (_, row) in enumerate(df_source.iterrows(), start=2):
                value = row[col_name]
                # Gérer les valeurs NaN
                if pd.isna(value):
                    cell_value = None
                elif isinstance(value, (np.integer, np.int64, np.int32)):
                    cell_value = int(value)
                elif isinstance(value, (np.floating, np.float64, np.float32)):
                    cell_value = float(value)
                elif isinstance(value, np.bool_):
                    cell_value = bool(value)
                else:
                    cell_value = value
                ws.cell(row=i, column=wb_col, value=cell_value)

        # Nettoyer les cellules au‑delà de la dernière ligne du DataFrame
        max_row = len(df_source) + 1
        if ws.max_row > max_row:
            for row in range(max_row + 1, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).value = None

        # 4. Sauvegarder et fermer
        wb.save(self.source_file_path)
        wb.close()
        print(f"🎨 Fichier sauvegardé avec les données mises à jour.")

    # ------------------------------------------------------------------
    # Extraction des codes sites
    # ------------------------------------------------------------------
    def _extract_site_code(self, text: str) -> Optional[str]:
        if not isinstance(text, str):
            text = str(text) if text is not None else ""
        pattern = re.compile('|'.join(re.escape(p) for p in self.extract_prefixes))
        match = pattern.search(text)
        if match:
            start = match.start()
            # On suppose que le code fait 7 caractères (à ajuster si besoin)
            return text[start:start+7]
        return None

    def _prepare_target_with_key(self, df_target: pd.DataFrame) -> tuple[pd.DataFrame, str]:
        # 1. target_key_column explicite
        if self.target_key_column is not None:
            key_col = self._resolve_column_name(df_target, self.target_key_column)
            missing = df_target[key_col].isna().sum()
            if missing > 0:
                raise ValueError(f"La colonne clé '{key_col}' contient {missing} valeurs manquantes.")
            print(f"🔑 Utilisation de la colonne clé existante : '{key_col}'")
            return df_target, key_col

        # 2. Extraction depuis extract_source_column
        if self.extract_source_column is not None:
            source_col = self._resolve_column_name(df_target, self.extract_source_column)
            print(f"🔍 Extraction des codes sites depuis la colonne '{source_col}'")
            df_target = df_target.copy()
            df_target['__extracted_key__'] = df_target[source_col].astype(str).apply(self._extract_site_code)
            if df_target['__extracted_key__'].isna().all():
                print("⚠️ Aucun code site extrait (tous les champs sont vides ou sans préfixe).")
            return df_target, '__extracted_key__'

        # 3. Recherche par noms par défaut
        default_names = ["code site", "site code", "site id", "Site Code", "Code Site", "siteID", "SiteID"]
        try:
            key_col = self._find_column_by_names(df_target, default_names)
            print(f"🔑 Colonne clé trouvée automatiquement : '{key_col}'")
            missing = df_target[key_col].isna().sum()
            if missing > 0:
                print(f"⚠️ La colonne '{key_col}' contient {missing} valeurs manquantes.")
            return df_target, key_col
        except KeyError:
            raise ValueError(
                "Impossible de trouver une colonne clé. Veuillez spécifier 'target_key_column' "
                "ou 'extract_source_column' dans les paramètres."
            )

    def _build_joined_value(self, row: pd.Series, columns: List[str]) -> str:
        values = []
        for col in columns:
            val = row.get(col, "")
            if pd.notna(val):
                str_val = str(val).strip()
                if str_val or not self.ignore_empty:
                    values.append(str_val)
        return self.join_separator.join(values)

    # ------------------------------------------------------------------
    # Méthode principale
    # ------------------------------------------------------------------
    def run(self) -> None:
        # Initialiser la date si non fournie
        if self.reference_date is None:
            self.reference_date = datetime.now().strftime("%Y-%m-%d")

        print("🔍 Démarrage XLOOKUP avec extraction des codes sites...")

        self._remove_auto_filters(self.source_file_path, self.source_sheet_name or None)
        self._remove_auto_filters(self.target_file_path, self.target_sheet_name or None)

        df_source = pd.read_excel(
            self.source_file_path,
            sheet_name=self.source_sheet_name or 0,
            engine='openpyxl'
        )
        df_target = pd.read_excel(
            self.target_file_path,
            sheet_name=self.target_sheet_name or 0,
            engine='openpyxl'
        )

        source_key_col = self._resolve_column_name(df_source, self.source_key_column)
        df_source[source_key_col] = df_source[source_key_col].astype(str).str.strip().replace("nan", "").fillna("")

        df_target_ready, target_key_col = self._prepare_target_with_key(df_target)
        df_target_ready[target_key_col] = df_target_ready[target_key_col].astype(str).str.strip().replace("nan", "").fillna("")

        # Construire la valeur à rapporter (TEXTJOIN ou colonne unique)
        if self.target_join_columns is not None:
            join_cols_real = self._resolve_columns_list(df_target_ready, self.target_join_columns)
            df_target_ready['__joined_value__'] = df_target_ready.apply(
                lambda row: self._build_joined_value(row, join_cols_real), axis=1
            )
            value_col = '__joined_value__'
            print(f"🔗 Valeur construite par TEXTJOIN sur les colonnes : {join_cols_real}, séparateur='{self.join_separator}'")
        else:
            value_col = self._resolve_column_name(df_target_ready, self.target_value_column)
            print(f"📌 Utilisation de la colonne unique : {value_col}")

        # Créer le mapping (garder la première occurrence si clés en double)
        df_unique = df_target_ready.drop_duplicates(subset=[target_key_col], keep='first')
        mapping = dict(zip(df_unique[target_key_col], df_unique[value_col]))

        # Déterminer la position de la colonne résultat
        result_idx = self._determine_result_column_position(df_source)
        result_col_name = self.result_column_name.strip()

        # Si la colonne résultat existe déjà, on la supprime pour la recréer à la position voulue
        if result_col_name in df_source.columns:
            df_source = df_source.drop(columns=[result_col_name])
            # Recalculer l'index après suppression
            result_idx = self._determine_result_column_position(df_source)

        # Insérer la nouvelle colonne à la position calculée
        if result_idx > len(df_source.columns):
            # Ajouter des colonnes vides si nécessaire
            for i in range(result_idx - len(df_source.columns)):
                df_source[f"_tmp_{i}"] = ""
            result_idx = len(df_source.columns)  # on peut maintenant insérer à la fin
        df_source.insert(result_idx, result_col_name, "")

        # Remplir les valeurs
        df_source[result_col_name] = df_source[source_key_col].map(mapping).fillna("")

        # Ajouter le préfixe
        if self.reference_name:
            prefix = f"{{{self.reference_name} - {self.reference_date}}} : "
            def add_prefix(cell):
                if pd.notna(cell) and str(cell).strip() != "":
                    return f"{prefix}{cell}"
                return ""
            df_source[result_col_name] = df_source[result_col_name].apply(add_prefix)

        # Supprimer les éventuelles colonnes temporaires ajoutées pour le décalage
        temp_cols = [col for col in df_source.columns if col.startswith("_tmp_")]
        if temp_cols:
            df_source = df_source.drop(columns=temp_cols)

        # Sauvegarder avec la méthode corrigée
        self._save_with_formatting(df_source)
        print(f"✅ Fichier source sauvegardé avec succès : {self.source_file_path}")
        print(f"📊 Colonne résultat : {result_col_name}")


# =============================================================================
# Exemple d'utilisation
# =============================================================================

if __name__ == "__main__":
    hourlyIHS = HourlyIHS(
        source_file_path=r"C:\Users\f50056342\Desktop\computer science\NUR Project Lyne\inputs\Book1.xlsx",
        target_file_path=r"C:\Users\f50056342\Desktop\emails_export\Hourly IHS.xlsx",
        source_key_column=["code site", "site code", "site id"],
        result_position_column="last_free",
        result_column_name="Hourly IHS",
        source_sheet_name="Sheet1",
        target_sheet_name="ALL",
        reference_name="Hourly IHS",
        reference_date=None,  # sera automatiquement la date du jour

        # --- Utilisation de NOMS de colonnes pour le TEXTJOIN ---
        target_join_columns=[
            "Event time",
            "Duration",
            "CMS",
            "Last Timeline Message",
            "Short description"
        ],
        join_separator=" .. ",
        ignore_empty=True,

        # --- Extraction automatique de la clé depuis la colonne "T" (ou nom) ---
        # extract_source_column="T",              # ou ["site code", "code site"]
        target_key_column=["Short description", "Duration","Outage Start Time"]  # si vous avez déjà une colonne clé
    )
    hourlyIHS.run()

    # Affichage du résultat
    df = pd.read_excel(r"C:\Users\f50056342\Desktop\computer science\NUR Project Lyne\inputs\Book1.xlsx")
    print("\n📄 Aperçu du fichier source après exécution :")
    print(df.head(30))