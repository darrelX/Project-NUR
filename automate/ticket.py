import os
from pathlib import Path
from typing import List, Optional, Any, Union
from dataclasses import dataclass, field
import pandas as pd
from datetime import datetime
import re
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows



@dataclass
class Ticket:
    """
    Classe pour manipuler des fichiers Excel :
    - Extraction automatique de codes sites (préfixes définis) dans une colonne cible.
    - XLOOKUP entre source et cible.
    - Possibilité de concaténer plusieurs colonnes cibles (TEXTJOIN) au lieu d'une seule.
    - Ajout d'un préfixe [reference_name - date] devant les résultats.
    - Suppression automatique des filtres Excel pour éviter les bugs.
    """

    source_file_path: str
    target_file_path: str
    source_key_column: Union[str, List[str]]  # colonne clé dans la source (nom(s) ou lettre)
    result_position_column: str               # colonne où écrire le résultat
    result_column_name: str                  # nom de la colonne résultat
    source_sheet_name: str = ""              # feuille source (défaut = première)
    target_sheet_name: str = ""              # feuille cible (défaut = première)
    reference_name: str = ""                 # pour le préfixe [NOM - date]
    reference_date: str = ""                 # date de référence (format : "DD/MM/YYYY")

    # Paramètres pour la valeur à rapporter depuis la cible
    target_value_column: Optional[Union[str, List[str]]] = None   # colonne unique (optionnel si join utilisée)
    target_join_columns: Optional[List[Union[str, List[str]]]] = None  # liste des colonnes à concaténer
    join_separator: str = ".."                # séparateur pour TEXTJOIN
    ignore_empty: bool = True                 # ignorer les cellules vides (comportement TEXTJOIN)

    # Paramètres pour l'extraction automatique de la clé cible
    target_key_column: Optional[Union[str, List[str]]] = None   # colonne clé cible (si déjà existante)
    extract_source_column: Optional[Union[str, List[str]]] = None  # colonne où chercher les codes
    extract_prefixes: Optional[List[str]] = None  # liste des préfixes (défaut ci-dessous)

    _source_path: Path = field(init=False, repr=False)
    _target_path: Path = field(init=False, repr=False)

    def __post_init__(self):
        self._source_path = Path(self.source_file_path)
        self._target_path = Path(self.target_file_path)

        # Date de référence par défaut = date du jour
        if not self.reference_date:
            self.reference_date = datetime.now().strftime("%d/%m/%Y")

        # Préfixes par défaut
        if self.extract_prefixes is None:
            self.extract_prefixes = [
                "CTR_", "SUO_", "SUD_", "NRO_", "ADM_",
                "NRD_", "EXN_", "LIT_", "EST_", "OST_"
            ]

        # Vérification qu'au moins une méthode de valeur cible est fournie
        if self.target_value_column is None and self.target_join_columns is None:
            raise ValueError("Vous devez spécifier soit target_value_column, soit target_join_columns.")

    # ------------------------------------------------------------------
    # Suppression des filtres Excel (AutoFilter)
    # ------------------------------------------------------------------
    @staticmethod
    def _remove_auto_filters(file_path: str, sheet_name: Optional[str] = None) -> None:
        """
        Supprime l'AutoFilter d'une seule feuille Excel.
        - sheet_name : nom de la feuille à nettoyer.
        - Si sheet_name est None, on utilise la première feuille du classeur.
        """
        try:
            wb = load_workbook(file_path)
            # Déterminer la feuille à traiter
            if sheet_name is None:
                target_sheet = wb.worksheets[0]          # première feuille
            else:
                target_sheet = wb[sheet_name]            # feuille par nom
            if target_sheet.auto_filter:
                target_sheet.auto_filter = None
                wb.save(file_path)
                print(f"🧹 Filtre supprimé sur la feuille '{target_sheet.title}' de {file_path}")
        except Exception as e:
            print(f"⚠️ Impossible de supprimer le filtre de {file_path} (feuille {sheet_name}) : {e}")

    # ------------------------------------------------------------------
    # Utilitaires de conversion lettres Excel ↔ indices
    # ------------------------------------------------------------------
    @staticmethod
    def _col_letter_to_index(col: str) -> int:
        col = col.upper().strip()
        result = 0
        for ch in col:
            result = result * 26 + (ord(ch) - ord('A') + 1)
        return result - 1

    @staticmethod
    def _index_to_col_letter(idx: int) -> str:
        result = ""
        idx += 1
        while idx > 0:
            idx, remainder = divmod(idx - 1, 26)
            result = chr(65 + remainder) + result
        return result

    def _find_column_by_names(self, df: pd.DataFrame, column_names: Union[str, List[str]]) -> str:
        """
        Trouve une colonne par nom(s), insensible à la casse.
        Si une liste est fournie, retourne la première colonne trouvée.
        """
        # Convertir en liste si c'est une chaîne
        if isinstance(column_names, str):
            column_names = [column_names]
        
        # Créer un mapping nom_lowercase -> nom_réel
        col_map = {str(col).strip().lower(): col for col in df.columns}
        
        # Chercher la première correspondance
        for name in column_names:
            name_lower = str(name).strip().lower()
            if name_lower in col_map:
                return col_map[name_lower]
        
        # Aucune correspondance trouvée
        raise KeyError(f"Aucune des colonnes {column_names} n'a été trouvée")
    
    def _resolve_column_name(self, df: pd.DataFrame, column_spec: Union[str, List[str]]) -> str:
        """
        Retourne le nom réel de la colonne (en-tête) à partir d'une spécification.
        Supporte : liste de noms, nom unique, numéro (1-indexé), lettre Excel.
        Insensible à la casse pour les noms.
        """
        # Si c'est une liste, chercher par noms
        if isinstance(column_spec, list):
            return self._find_column_by_names(df, column_spec)
        
        spec = str(column_spec).strip()
        
        # 1. Vérifier si c'est un index numérique
        if spec.isdigit():
            idx = int(spec) - 1
            if 0 <= idx < len(df.columns):
                return df.columns[idx]
            raise IndexError(f"Colonne numéro {spec} hors limites")
        
        # 2. Chercher par nom (insensible à la casse)
        try:
            return self._find_column_by_names(df, spec)
        except KeyError:
            pass
        
        # 3. Vérifier si c'est une lettre de colonne (A, B, AA...)
        # Limiter à 1-2 caractères pour éviter confusion avec noms
        if spec.isalpha() and len(spec) <= 2:
            idx = self._col_letter_to_index(spec)
            if 0 <= idx < len(df.columns):
                return df.columns[idx]
            raise IndexError(f"Colonne lettre {spec} hors limites")
        
        # 4. Si rien ne correspond
        raise KeyError(f"Colonne '{spec}' introuvable")

    def _resolve_columns_list(self, df: pd.DataFrame, columns_spec: List[str]) -> List[str]:
        """Convertit une liste de spécifications de colonnes (lettres/noms) en noms réels."""
        return [self._resolve_column_name(df, col) for col in columns_spec]

    def _determine_result_column_position(self, df_source: pd.DataFrame) -> int:
        """
        Retourne l'index (0-based) de la colonne résultat.
        Gère 'last_column' pour ajouter à la fin.
        """
        pos = str(self.result_position_column).strip().lower()
        if pos in ("last_free", "derniere_libre", "dernière_libre", "last_column", "derniere_colonne", "dernière_colonne"):
            return len(df_source.columns)
        if pos.isdigit():
            return int(pos) - 1
        if pos.isalpha():
            return self._col_letter_to_index(pos)
        if pos in df_source.columns:
            return df_source.columns.get_loc(pos)
        raise ValueError(f"Position de colonne invalide : '{self.result_position_column}'")

    def _save_with_formatting(self, df_source: pd.DataFrame) -> None:
        """
        Ajoute uniquement les NOUVELLES colonnes à droite du fichier Excel.
        Conserve TOUTES les données existantes sans y toucher.
        """
        import numpy as np
        
        # Charger le workbook existant
        wb = load_workbook(self.source_file_path)
        
        # Déterminer la feuille à modifier
        if isinstance(self.source_sheet_name, str) and self.source_sheet_name:
            ws = wb[self.source_sheet_name]
        else:
            ws = wb.active
        
        # Identifier les colonnes existantes dans Excel
        existing_cols = []
        for col_idx in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col_idx).value
            if header:
                existing_cols.append(header)
        
        # Identifier les NOUVELLES colonnes (celles à ajouter)
        new_columns = []
        new_col_indices = []
        for idx, col_name in enumerate(df_source.columns):
            if col_name not in existing_cols:
                new_columns.append(col_name)
                new_col_indices.append(idx)
        
        if not new_columns:
            print("✓ Aucune nouvelle colonne à ajouter")
            wb.close()
            return
        
        # Ajouter les en-têtes des nouvelles colonnes à droite
        start_col = len(existing_cols) + 1
        for i, col_name in enumerate(new_columns):
            ws.cell(row=1, column=start_col + i, value=col_name)
        
        # Écrire UNIQUEMENT les données des nouvelles colonnes
        for row_idx, row_data in enumerate(df_source.values, start=2):
            for i, df_col_idx in enumerate(new_col_indices):
                value = row_data[df_col_idx]
                
                # Convertir les types pandas/numpy en types Python natifs
                if pd.isna(value):
                    cell_value = None
                elif isinstance(value, (np.integer, np.int64, np.int32)):
                    cell_value = int(value)
                elif isinstance(value, (np.floating, np.float64, np.float32)):
                    cell_value = float(value)
                elif isinstance(value, np.bool_):
                    cell_value = bool(value)
                else:
                    cell_value = value
                
                ws.cell(row=row_idx, column=start_col + i, value=cell_value)
        
        # Sauvegarder et fermer
        wb.save(self.source_file_path)
        wb.close()
        print(f"🎨 {len(new_columns)} nouvelle(s) colonne(s) ajoutée(s) : {', '.join(new_columns)}")

    # ------------------------------------------------------------------
    # Extraction des codes sites (équivalent à la formule Excel)
    # ------------------------------------------------------------------
    def _extract_site_code(self, text: str) -> Optional[str]:
        """
        Extrait le code site selon la logique de la formule Excel.
        Utilise une regex pour trouver l'un des préfixes, puis renvoie les 7 caractères
        (préfixe + 3 caractères suivants, ou moins si la chaîne est trop courte).
        Retourne None si aucun code n'est trouvé.
        """
        if not isinstance(text, str):
            text = str(text) if text is not None else ""

        # Construction du pattern regex
        pattern = re.compile('|'.join(re.escape(p) for p in self.extract_prefixes))
        match = pattern.search(text)
        if match:
            start = match.start()
            extracted = text[start:start+7]
            return extracted
        return None

    def _prepare_target_with_key(self, df_target: pd.DataFrame) -> tuple[pd.DataFrame, str]:
        """
        Prépare le DataFrame cible en lui ajoutant une colonne 'site_code' (si target_key_column non fourni).
        Si target_key_column est fourni et existe, on l'utilise directement.
        Sinon, on crée la colonne à partir de extract_source_column.
        
        Returns:
            tuple (df_target, key_column_name)
        """
        if self.target_key_column:
            key_col = self._resolve_column_name(df_target, self.target_key_column)
            missing_keys = df_target[key_col].isna().sum()
            if missing_keys > 0:
                raise ValueError(
                    f"La colonne clé '{key_col}' contient {missing_keys} valeurs manquantes (NaN)."
                )
            return df_target, key_col
        else:
            # Chercher automatiquement une colonne contenant "site id" (insensible à la casse)
            matching_cols = [col for col in df_target.columns if 'site id' in col.lower()]
            if not matching_cols:
                raise ValueError(
                    "Ni target_key_column ni extract_source_column n'est défini, "
                    "et aucune colonne contenant 'site id' n'a été trouvée."
                )
            source_col = matching_cols[0]
            df_target = df_target.copy()
            df_target['__extracted_key__'] = df_target[source_col].astype(str).apply(self._extract_site_code)
            return df_target, '__extracted_key__'

    def _build_joined_value(self, row: pd.Series, columns: List[str]) -> str:
        """
        Construit la chaîne TEXTJOIN à partir des colonnes spécifiées.
        Similaire à : =TEXTJOIN(separator, ignore_empty, plage)
        """
        values = []
        for col in columns:
            val = row.get(col, "")
            if pd.notna(val):
                str_val = str(val).strip()
                if str_val or not self.ignore_empty:
                    values.append(str_val)
        return self.join_separator.join(values)

    # ------------------------------------------------------------------
    # Méthode principale
    # ------------------------------------------------------------------
    def run(self) -> None:
        """Exécute le XLOOKUP avec extraction automatique et option TEXTJOIN."""
        print("🔍 Démarrage XLOOKUP avec extraction des codes sites...")

        self._remove_auto_filters(self.source_file_path, self.source_sheet_name or None)
        self._remove_auto_filters(self.target_file_path, self.target_sheet_name or None)

        # 2. Chargement source et cible (forcer engine='openpyxl' pour cohérence)
        df_source = pd.read_excel(
            self.source_file_path,
            sheet_name=self.source_sheet_name or 0,
            engine='openpyxl'
        )
        df_target = pd.read_excel(
            self.target_file_path,
            sheet_name=self.target_sheet_name or 0,
            engine='openpyxl'
        )

        # Résolution colonne clé source
        source_key_col = self._resolve_column_name(df_source, self.source_key_column)
        df_source[source_key_col] = df_source[source_key_col].astype(str).str.strip()

        # Préparation de la cible : création de la colonne clé si besoin
        df_target_ready, target_key_col = self._prepare_target_with_key(df_target)
        df_target_ready[target_key_col] = df_target_ready[target_key_col].astype(str).str.strip()

        # Construction de la colonne de valeurs à mapper (soit colonne unique, soit TEXTJOIN)
        if self.target_join_columns is not None:
            # TEXTJOIN sur plusieurs colonnes
            join_cols_real = self._resolve_columns_list(df_target_ready, self.target_join_columns)
            df_target_ready['__joined_value__'] = df_target_ready.apply(
                lambda row: self._build_joined_value(row, join_cols_real), axis=1
            )
            value_col = '__joined_value__'
            print(f"🔗 Valeur construite par TEXTJOIN sur les colonnes : {join_cols_real}, séparateur='{self.join_separator}'")
        else:
            # Colonne unique
            value_col = self._resolve_column_name(df_target_ready, self.target_value_column)
            print(f"📌 Utilisation de la colonne unique : {value_col}")

        # Mapping clé -> valeur
        mapping = dict(zip(df_target_ready[target_key_col], df_target_ready[value_col]))

        # Détermination de la colonne résultat dans la source
        result_idx = self._determine_result_column_position(df_source)
        result_col_name = self.result_column_name.strip()

        # Supprimer toutes les colonnes existantes avec ce nom (éviter les doublons)
        if result_col_name in df_source.columns:
            df_source = df_source.loc[:, ~df_source.columns.duplicated()]
            cols_to_drop = [col for col in df_source.columns if col == result_col_name]
            df_source = df_source.drop(columns=cols_to_drop)
            # Recalculer result_idx après suppression
            result_idx = self._determine_result_column_position(df_source)

        # Ajout ou renommage de la colonne résultat
        if result_idx >= len(df_source.columns):
            for _ in range(result_idx - len(df_source.columns) + 1):
                df_source[f"_tmp_{_}"] = ""
            df_source.rename(columns={df_source.columns[result_idx]: result_col_name}, inplace=True)
        else:
            existing = df_source.columns[result_idx]
            if existing != result_col_name:
                df_source.rename(columns={existing: result_col_name}, inplace=True)

        # Application du lookup
        df_source[result_col_name] = df_source[source_key_col].map(mapping).fillna("")

        if self.reference_name:
            prefix = f"{{{self.reference_name} - {self.reference_date}}} : "

            def add_prefix(cell):
                if pd.notna(cell) and str(cell).strip() != "":
                    return f"{prefix}{cell}"
                return ""

            # Sélection de la/des colonne(s)
            selected = df_source[result_col_name]

            print(f"Type de selected : {type(selected)}")
            if isinstance(selected, pd.DataFrame):
                print(f"selected est un DataFrame avec {len(selected.columns)} colonnes : {list(selected.columns)}")
                # On prend la première colonne (on suppose que c'est celle qui nous intéresse)
                selected = selected.iloc[:, 0]  # .iloc[:, 0] donne la première colonne en Series
                print("On utilise la première colonne")
            else:
                print("selected est une Series")

            # Maintenant selected est forcément une Series
            df_source[result_col_name] = selected.apply(add_prefix)

        # Sauvegarde du fichier modifié avec conservation de la mise en forme
        self._save_with_formatting(df_source)
        print(f"✅ Fichier source sauvegardé avec succès : {self.source_file_path}")
        print(f"📊 Colonne résultat : {result_col_name}")


# =============================================================================
# Exemple d'utilisation avec TEXTJOIN sur les colonnes L, W, X, Y, Z
# =============================================================================
if __name__ == "__main__":
    ticket = Ticket(
        source_file_path=r"C:\Users\f50056342\Desktop\computer science\NUR Project Lyne\inputs\Book1.xlsx",
        target_file_path=r"C:\Users\f50056342\Downloads\Incident Ticket_20260524232156.xlsx",
        source_key_column="B",                 # Codesite dans colonne B
        result_position_column="last_free",            # Résultat en colonne C
        result_column_name="Ticket",
        source_sheet_name="Sheet1",
        target_sheet_name="",                  # première feuille du cible
        reference_name="Ticket",
        # --- Utilisation du TEXTJOIN ---
        target_join_columns=["L", "W", "X", "Y", "Z"],   # colonnes à concaténer
        join_separator="..",                    # séparateur
        ignore_empty=True,                      # ignorer les vides
        # --- Extraction de la clé cible ---
        extract_source_column="T",              # colonne contenant les préfixes
        # target_key_column=None (extraction auto)
    )
    ticket.run()

    # Affichage du résultat
    df = pd.read_excel(r"C:\Users\f50056342\Desktop\computer science\NUR Project Lyne\inputs\Book1.xlsx")
    print("\n📄 Aperçu du fichier source après exécution :")
    print(df.head(30))