# ============================================================================
# TÉLÉCHARGEMENT - À remplacer dans app.py
# Utilise openpyxl data_only=False pour préserver formules + couleurs
# Les nouvelles colonnes sont insérées après la colonne "impacts"
# ============================================================================

from io import BytesIO
import openpyxl

st.divider()
st.subheader("📥 Télécharger le fichier compilé")

# ── Ouvrir le fichier original avec openpyxl (formules + styles préservés) ──
wb = openpyxl.load_workbook(
    BytesIO(st.session_state.fichiers["principal_binaire"]),
    data_only=False   # False = garde les formules, True = garde les valeurs
)
ws = wb["daily break down"]

# ── Trouver la colonne "impacts" dans la ligne d'en-tête ────────────────────
col_impacts_idx = None  # index 1-based (openpyxl commence à 1)
for cell in ws[1]:      # ws[1] = première ligne = en-têtes
    if cell.value and str(cell.value).strip().lower() == "impacts":
        col_impacts_idx = cell.column  # ex: 20 si impacts est en colonne T
        break

if col_impacts_idx is None:
    st.error("❌ Colonne 'impacts' introuvable dans le fichier principal.")
    st.stop()

# ── Identifier les nouvelles colonnes à insérer ─────────────────────────────
# Ce sont toutes les colonnes de df_principal qui n'existaient pas à l'origine
colonnes_originales = [
    "Nom du Physique", "Codesite", "CLUSTER", "Owner", "Topology",
    "Region", "Ville", "Equipementiers", "NUR 2G", "NUR 3G", "NUR 4G",
    "NUR  DU JOUR", "oprationnel", "Raison d'exclusion",
    "OCM NOMENCLATURE", "CAUSE", "SUB RCA", "CATEGORIE",
    "Verifications", "impacts"
]
nouvelles_colonnes = [
    c for c in df_principal.columns
    if c not in colonnes_originales
]

if not nouvelles_colonnes:
    st.warning("⚠️ Aucune nouvelle colonne à ajouter — fichier original retourné tel quel.")
else:
    # ── Insérer les nouvelles colonnes juste après "impacts" ────────────────
    # openpyxl insert_cols décale tout ce qui est à droite
    nb_nouvelles = len(nouvelles_colonnes)
    position_insertion = col_impacts_idx + 1  # juste après impacts

    ws.insert_cols(position_insertion, nb_nouvelles)

    # ── Écrire les en-têtes des nouvelles colonnes ──────────────────────────
    for i, nom_col in enumerate(nouvelles_colonnes):
        ws.cell(row=1, column=position_insertion + i, value=nom_col)

    # ── Trouver la colonne Code site dans df_principal pour l'alignement ────
    col_code = None
    for col in df_principal.columns:
        if "codesite" in str(col).lower().replace(" ", ""):
            col_code = col
            break

    # ── Écrire les valeurs ligne par ligne ───────────────────────────────────
    # On aligne par Codesite entre df_principal et le fichier openpyxl
    # pour éviter tout décalage si les lignes ne sont pas dans le même ordre

    # Construire un dictionnaire codesite → valeurs nouvelles colonnes
    dict_valeurs = {}
    for _, row in df_principal.iterrows():
        code = str(row[col_code]).strip() if col_code else None
        dict_valeurs[code] = {c: row[c] for c in nouvelles_colonnes}

    # Trouver l'index de la colonne Codesite dans openpyxl
    col_codesite_idx = None
    for cell in ws[1]:
        if cell.value and "codesite" in str(cell.value).strip().lower().replace(" ", ""):
            col_codesite_idx = cell.column
            break

    # Écrire les valeurs dans les nouvelles colonnes
    for row_idx in range(2, ws.max_row + 1):  # ligne 2 = première ligne de données
        if col_codesite_idx:
            code_cell = ws.cell(row=row_idx, column=col_codesite_idx)
            code = str(code_cell.value).strip() if code_cell.value else ""
            valeurs_ligne = dict_valeurs.get(code, {})
        else:
            valeurs_ligne = {}

        for i, nom_col in enumerate(nouvelles_colonnes):
            valeur = valeurs_ligne.get(nom_col, "")
            # Convertir NaN en chaîne vide
            if isinstance(valeur, float) and str(valeur) == "nan":
                valeur = ""
            ws.cell(row=row_idx, column=position_insertion + i, value=valeur)

# ── Sauvegarder dans un buffer mémoire ──────────────────────────────────────
output = BytesIO()
wb.save(output)
output.seek(0)

# ── Bouton de téléchargement ─────────────────────────────────────────────────
st.download_button(
    label="📥 Télécharger le fichier compilé",
    data=output,
    file_name="compilation_noc_finale.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    type="primary",
    use_container_width=True
)