# ============================================================================
# app.py - Orchestrateur principal du Compilateur NOC
# ============================================================================

print("Heelo World")
import streamlit as st

# set_page_config DOIT être la première commande Streamlit, avant tout import
st.set_page_config(page_title="Compilateur NOC", layout="wide")

# Imports après set_page_config pour éviter ScriptRunContext warning
import pandas as pd
from io import BytesIO
import time
import openpyxl
from components.file_uploader import render_file_uploader
from config import SOURCES
st.title("🔧 Compilateur NOC")
st.caption("Chargez vos fichiers sources, configurez chaque module, puis lancez la compilation.")

# ============================================================================
# ÉTAPE 1 : CHARGEMENT DES FICHIERS
# ============================================================================
principal_ok = render_file_uploader()

# ============================================================================
# ÉTAPE 2 : COMPILATION
# ============================================================================
if principal_ok and st.button("🚀 Lancer la compilation", type="primary", use_container_width=True):

    start_time = time.time()

    df_principal = st.session_state.fichiers["principal"].copy()
    date_ref     = st.session_state.get("date_reference", "")

    # Trouver la colonne Code site
    col_code = None
    for col in df_principal.columns:
        if "code site" in str(col).lower() or "codesite" in str(col).lower():
            col_code = col
            break

    if col_code is None:
        st.error("❌ Colonne 'Code site' introuvable dans le fichier principal.")
        st.stop()

    total_sites       = len(df_principal)
    progress_container = st.container()

    with progress_container:
        st.subheader("📊 Progression de la compilation")
        st.write("---")

    # =========================================================================
    # 1. INCIDENT
    # =========================================================================
    if st.session_state.fichiers["sources"].get("incident"):
        with st.spinner("Compilation INCIDENT..."):
            from modules.incident import IncidentModule
            module_incident = IncidentModule()
            df_principal = module_incident.compiler(
                df_principal,
                st.session_state.fichiers["sources"]["incident"],
                col_code,
                date_ref
            )
        with progress_container:
            st.success("✅ INCIDENT compilé")

    # =========================================================================
    # 2. TOP OFFENDERS
    # =========================================================================
    if st.session_state.fichiers["sources"].get("top_offenders"):
        with st.spinner("Compilation Top Offenders..."):
            from modules.top_offenders import TopOffendersModule
            module_top = TopOffendersModule()
            df_principal = module_top.compiler(
                df_principal,
                st.session_state.fichiers["sources"]["top_offenders"],
                col_code,
                date_ref
            )
        with progress_container:
            st.success("✅ Top Offenders compilé")

    # =========================================================================
    # 3. RETOUR CAMUSAT
    # =========================================================================
    if st.session_state.fichiers["sources"].get("retour_camusat"):
        with st.spinner("Compilation RETOUR CAMUSAT..."):
            from modules.retour_camusat import RetourCamusatModule
            module_rc = RetourCamusatModule()
            df_principal = module_rc.compiler(
                df_principal,
                st.session_state.fichiers["sources"]["retour_camusat"],
                col_code,
                date_ref
            )
        with progress_container:
            st.success("✅ RETOUR CAMUSAT compilé")

    # =========================================================================
    # 4. RETOUR IHS
    # =========================================================================
    if st.session_state.fichiers["sources"].get("retour_ihs"):
        with st.spinner("Compilation RETOUR IHS..."):
            from modules.retour_ihs import RetourIhsModule
            module_ri = RetourIhsModule()
            df_principal = module_ri.compiler(
                df_principal,
                st.session_state.fichiers["sources"]["retour_ihs"],
                col_code,
                date_ref
            )
        with progress_container:
            st.success("✅ RETOUR IHS compilé")

    # =========================================================================
    # 5. HOURLY IHS
    # =========================================================================
    if st.session_state.fichiers["sources"].get("hourly_ihs"):
        with st.spinner("Compilation Hourly IHS..."):
            from modules.hourly_ihs import HourlyIhsModule
            module_hi = HourlyIhsModule()
            df_principal = module_hi.compiler(
                df_principal,
                st.session_state.fichiers["sources"]["hourly_ihs"],
                col_code
            )
        with progress_container:
            st.success("✅ Hourly IHS compilé")

    # =========================================================================
    # 6. CELLS DOWN
    # =========================================================================
    if st.session_state.fichiers["sources"].get("cells_down"):
        with st.spinner("Compilation Cells DOWN..."):
            from modules.cells_down import CellsDownModule
            module_cells = CellsDownModule()
            df_principal = module_cells.compiler(
                df_principal,
                st.session_state.fichiers["sources"]["cells_down"],
                col_code,
                date_ref
            )
        with progress_container:
            st.success("✅ Cells DOWN compilé")

    # =========================================================================
    # 7. OCM RAN
    # =========================================================================
    if st.session_state.fichiers["sources"].get("ocm_ran"):
        with st.spinner("Compilation OCM RAN..."):
            from modules.ocm_ran_incident import OcmRanModule
            module_ocm = OcmRanModule()
            df_principal = module_ocm.compiler(
                df_principal,
                st.session_state.fichiers["sources"]["ocm_ran"],
                col_code,
                date_ref
            )
        with progress_container:
            st.success("✅ OCM RAN compilé")

    # =========================================================================
    # 8. TICKET
    # =========================================================================
    if st.session_state.fichiers["sources"].get("ticket"):
        with st.spinner("Compilation Ticket..."):
            from modules.ticket import TicketModule
            module_ticket = TicketModule()
            df_principal = module_ticket.compiler(
                df_principal,
                st.session_state.fichiers["sources"]["ticket"],
                col_code,
                date_ref
            )
        with progress_container:
            st.success("✅ Ticket compilé")
            
            # Ajouter téléchargement du fichier ticket modifié (COMPLET)
            ticket_data, ticket_filename = module_ticket.exporter_ticket()
            if ticket_data:
                st.download_button(
                    label=f"📥 Télécharger {ticket_filename}",
                    data=ticket_data,
                    file_name=ticket_filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="download_ticket"
                )

    # =========================================================================
    # 9. DASHBOARD CELL
    # =========================================================================
    if st.session_state.fichiers["sources"].get("dashboard_cell"):
        with st.spinner("Compilation Dashboard Cell..."):
            from modules.dashboard_cell import DashboardCellModule
            module_dash = DashboardCellModule()
            df_principal = module_dash.compiler(
                df_principal,
                st.session_state.fichiers["sources"]["dashboard_cell"],
                col_code,
                date_ref
            )
        with progress_container:
            st.success("✅ Dashboard Cell compilé")

    temps_total = time.time() - start_time

    # =========================================================================
    # STATISTIQUES
    # =========================================================================
    st.divider()
    st.subheader("📊 Résultat final")

    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("Total sites", total_sites)
    with col2:
        st.metric("Temps", f"{temps_total:.1f}s")
    with col3:
        colonnes_originales = [
            "Nom du Physique", "Codesite", "CLUSTER", "Owner", "Topology",
            "Region", "Ville", "Equipementiers", "NUR 2G", "NUR 3G", "NUR 4G",
            "NUR  DU JOUR", "oprationnel", "Raison d'exclusion",
            "OCM NOMENCLATURE", "CAUSE", "SUB RCA", "CATEGORIE",
            "Verifications", "impacts"
        ]
        nouvelles_colonnes = [c for c in df_principal.columns if c not in colonnes_originales]
        st.metric("Nouvelles colonnes", len(nouvelles_colonnes))

    # =========================================================================
    # APERÇU
    # =========================================================================
    st.subheader("📋 Aperçu du fichier compilé")
    if nouvelles_colonnes:
        st.info(f"🔍 {len(nouvelles_colonnes)} nouvelles colonnes ajoutées : {', '.join(nouvelles_colonnes)}")
    st.dataframe(df_principal, use_container_width=True)

    # =========================================================================
    # TÉLÉCHARGEMENT avec openpyxl
    # Préserve : formules, couleurs, styles du fichier original
    # Insère les nouvelles colonnes juste après "impacts"
    # =========================================================================
    st.divider()
    st.subheader("📥 Télécharger le fichier compilé")

    with st.spinner("Préparation du fichier final (préservation des formules)..."):

        # Ouvrir le fichier ORIGINAL avec openpyxl
        # data_only=False → garde les formules (=VLOOKUP, =IF, etc.)
        wb = openpyxl.load_workbook(
            BytesIO(st.session_state.fichiers["principal_binaire"]),
            data_only=False
        )
        ws = wb["daily break down"]

        # Trouver l'index de la colonne "impacts" (1-based)
        col_impacts_idx = None
        colonnes_disponibles = []
        for cell in ws[1]:
            if cell.value:
                col_name = str(cell.value).strip().lower()
                colonnes_disponibles.append(str(cell.value).strip())
                # Recherche robuste : ignorer espaces supplémentaires
                if "impact" in col_name:
                    col_impacts_idx = cell.column
                    break

        if col_impacts_idx is None:
            st.warning(f"⚠️ Colonne 'impacts' introuvable. Colonnes disponibles : {', '.join(colonnes_disponibles)}")
            # Si pas trouvée, utiliser la dernière colonne non-vide comme référence
            col_impacts_idx = ws.max_column

        if nouvelles_colonnes:
            # Insérer les nouvelles colonnes juste après "impacts"
            position_insertion = col_impacts_idx + 1
            ws.insert_cols(position_insertion, len(nouvelles_colonnes))

            # Écrire les en-têtes
            for i, nom_col in enumerate(nouvelles_colonnes):
                ws.cell(row=1, column=position_insertion + i, value=nom_col)

            # Construire dictionnaire codesite → nouvelles valeurs
            dict_valeurs = {}
            for _, row in df_principal.iterrows():
                code = str(row[col_code]).strip()
                dict_valeurs[code] = {c: row[c] for c in nouvelles_colonnes}

            # Trouver l'index Codesite dans openpyxl
            col_codesite_idx = None
            for cell in ws[1]:
                if cell.value and "codesite" in str(cell.value).strip().lower().replace(" ", ""):
                    col_codesite_idx = cell.column
                    break

            # Écrire les valeurs alignées par Codesite
            for row_idx in range(2, ws.max_row + 1):
                if col_codesite_idx:
                    code = str(ws.cell(row=row_idx, column=col_codesite_idx).value or "").strip()
                    valeurs_ligne = dict_valeurs.get(code, {})
                else:
                    valeurs_ligne = {}

                for i, nom_col in enumerate(nouvelles_colonnes):
                    valeur = valeurs_ligne.get(nom_col, "")
                    # Convertir NaN en cellule vide
                    if isinstance(valeur, float) and str(valeur) == "nan":
                        valeur = ""
                    ws.cell(row=row_idx, column=position_insertion + i, value=valeur)

        # Sauvegarder dans un buffer mémoire
        output = BytesIO()
        wb.save(output)
        output.seek(0)

    # =========================================================================
    # LIBÉRER LA RAM après compilation (problème 6)
    # Supprimer les binaires des fichiers sources du session_state
    # =========================================================================
    for key in st.session_state.fichiers["sources"]:
        st.session_state.fichiers["sources"][key] = []
    st.session_state.fichiers["principal"]         = None
    st.session_state.fichiers["principal_binaire"] = None
    st.session_state.fichiers["principal_nom"]     = None

    st.download_button(
        label="📥 Télécharger le fichier compilé",
        data=output,
        file_name="compilation_noc_finale.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        use_container_width=True
    )

else:
    if not principal_ok:
        st.info("📂 Veuillez charger le fichier principal pour commencer.")