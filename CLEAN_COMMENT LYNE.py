"""
=====================================================
NETTOYAGE DES COMMENTAIRES - BREAKDOWN
=====================================================
Utilise ce script pour nettoyer la colonne Verification
de n'importe quel fichier daily-break-down.

INSTRUCTIONS:
1. Configure le CHEMIN du fichier ci-dessous (ligne 20)
2. Execute: python COMMENT_CLEAN.py
=====================================================
"""

# ========================================================
# CONFIGURATION - MODIFIE ICI LE CHEMIN DE TON FICHIER
# ========================================================

CHEMIN_FICHIER = r"C:\Users\f50056342\Desktop\my work\breakdown daily\daily-break-down 06-14-2026.xlsx"

# ========================================================
# SCRIPT - NE PAS MODIFIER
# ========================================================

from openpyxl import load_workbook
from openpyxl.styles import Font
import re
import os

# Séparateur entre les blocs
SEP = "     ;     "

# Bruits à supprimer
NOISES = [
    r'Site\s*down', r'Site\s*Down',
    r'Auto\s*TT\s*closure\s*RCA',
    r'Power\s*status\s*check\s*(?:Ongoing)?',
    r'Others', r'other',
    r'Need\s+BO\s+RAN\s*analysis',
    r'Besoin\s+d\'analyse\s+BO\s+RAN',
    r'ANALYSE\s+BO\s+TX',
]


def clean_comment(comment):
    """Nettoie un commentaire selon les règles définies."""
    if not comment:
        return ""
    
    result = str(comment)
    
    # 1. Supprimer les tickets
    result = re.sub(r'INC-\d{8}-\d+', ' ', result)
    result = re.sub(r'CM-\d{8}-\d+', ' ', result)
    
    # 2. Supprimer les bruits avec date
    for noise in NOISES:
        result = re.sub(r'\d{1,2}/\d{1,2}/\d{4}\s*:\s*' + noise, ' ', result, flags=re.IGNORECASE)
    
    # 3. Supprimer les bruits sans date
    for noise in NOISES:
        result = re.sub(noise, ' ', result, flags=re.IGNORECASE)
    
    # 4. Supprimer les blocs {}
    result = re.sub(r'\{[^}]*\}\s*:\s*', '', result)
    result = re.sub(r'\{[^}]*\}', ' ', result)
    
    # 5. Remplacer les .. par le séparateur
    result = re.sub(r'\.\.\s*:\s*', SEP, result)
    result = re.sub(r'\s*\.\.\s*', SEP, result)
    result = re.sub(r'\s*;\s*', SEP, result)
    
    # 6. Garder les || intacts
    if '||' in result:
        parts = result.split('||')
        cleaned = []
        for part in parts:
            part = part.strip()
            if not part:
                continue
            part = re.sub(r'^\s*:\s*', '', part)
            for noise in NOISES:
                part = re.sub(noise, ' ', part, flags=re.IGNORECASE)
            part = part.strip()
            if part:
                cleaned.append(part)
        return ' || '.join(cleaned)
    
    # 7. Split et dédoublonner
    parts = [p.strip() for p in result.split(SEP)]
    cleaned_parts = []
    seen = set()
    
    for part in parts:
        part = re.sub(r'^\s*:\s*', '', part).strip()
        if not part or part.isspace():
            continue
        normalized = part.lower().strip()
        if normalized and normalized not in seen:
            seen.add(normalized)
            cleaned_parts.append(part)
    
    result = SEP.join(cleaned_parts)
    result = re.sub(r'\s+', ' ', result)
    result = result.strip()
    
    return result


def main():
    print("=" * 70)
    print("NETTOYAGE DES COMMENTAIRES - BREAKDOWN")
    print("=" * 70)
    
    # Vérifier le fichier
    if not os.path.exists(CHEMIN_FICHIER):
        print(f"\nERREUR: Fichier non trouve!")
        print(f"Chemin: {CHEMIN_FICHIER}")
        print("\nModifie la variable CHEMIN_FICHIER dans le script.")
        return
    
    print(f"\nFichier: {CHEMIN_FICHIER}")
    
    # Charger le fichier
    wb = load_workbook(CHEMIN_FICHIER)
    print(f"Sheets: {wb.sheetnames}")
    
    # Trouver la sheet principale (daily break down)
    sheet_name = None
    for name in wb.sheetnames:
        if 'daily break down' in name.lower():
            sheet_name = name
            if name.lower() == 'daily break down':
                break
    
    if sheet_name is None:
        sheet_name = wb.sheetnames[0]
    
    print(f"Sheet utilisee: {sheet_name}")
    ws = wb[sheet_name]
    
    # Trouver colonne Verification
    verification_col = None
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val and str(val).strip() == "Verification":
            verification_col = col
            break
    
    if verification_col is None:
        print("ERREUR: Colonne 'Verification' non trouvee!")
        return
    
    print(f"Colonne Verification: {verification_col}")
    
    # Créer cleaning_comment
    cleaning_col = verification_col + 1
    ws.cell(row=1, column=cleaning_col).value = "cleaning_comment"
    ws.cell(row=1, column=cleaning_col).font = Font(bold=True)
    print(f"Colonne cleaning_comment creee en: {cleaning_col}")
    
    # Traitement
    print("\n" + "-" * 70)
    print("TRAITEMENT...")
    print("-" * 70)
    
    processed = 0
    for row in range(2, ws.max_row + 1):
        original = ws.cell(row=row, column=verification_col).value
        if original:
            ws.cell(row=row, column=cleaning_col).value = clean_comment(original)
            processed += 1
    
    print(f"Lignes traitees: {processed}")
    
    # Exemples
    print("\nEXEMPLES:")
    count = 0
    for row in range(2, ws.max_row + 1):
        original = ws.cell(row=row, column=verification_col).value
        if original and count < 5:
            site = ws.cell(row=row, column=2).value
            cleaned = ws.cell(row=row, column=cleaning_col).value
            print(f"\n[{site}]")
            print(f"  AVANT: {str(original)[:80]}...")
            print(f"  APRES: {str(cleaned)[:80]}..." if cleaned else "  APRES: (vide)")
            count += 1
    
    # Sauvegarder
    base, ext = os.path.splitext(CHEMIN_FICHIER)
    output_path = f"{base}_CLEANED.xlsx"
    wb.save(output_path)
    
    print("\n" + "=" * 70)
    print("TERMINE!")
    print("=" * 70)
    print(f"\nFichier nettoye: {output_path}")
    
    # Vérification
    print("\nVerification:")
    wb2 = load_workbook(output_path)
    ws2 = wb2[sheet_name]
    print(f"  Header col {cleaning_col}: {ws2.cell(1, cleaning_col).value}")
    val2 = ws2.cell(2, cleaning_col).value
    print(f"  Row 2: {str(val2)[:50]}..." if val2 else "  Row 2: (vide)")


if __name__ == "__main__":
    main()
