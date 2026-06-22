# ==============================================================
# server.py - Serveur IA Breakdown (à exécuter sur le pod RunPod)
# ==============================================================
#
# Installation (sur le pod RunPod, dans un terminal) :
#   pip install fastapi uvicorn python-multipart pandas openpyxl
#   pip install transformers accelerate bitsandbytes torch
#
# Lancement :
#   python server.py
#   (ou : uvicorn server:app --host 0.0.0.0 --port 8000)
#
# ==============================================================

import os
import re
import time
import tempfile
import warnings
from typing import Optional, List

import torch
import pandas as pd
from openpyxl import load_workbook
from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.responses import FileResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from transformers import AutoTokenizer, AutoModelForCausalLM, BitsAndBytesConfig

warnings.filterwarnings("ignore")

# -- CONFIGURATION --------------------------------------------
# Adapte ce chemin à l'emplacement réel de model_final sur le pod
MODELE_PATH = "/workspace/model_final"

app = FastAPI(title="Breakdown IA - Serveur de classification")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],   # à restreindre à ton domaine en vraie prod
    allow_methods=["*"],
    allow_headers=["*"],
)

# -- 21 CATÉGORIES DE CLASSIFICATION ----------------------------
# Note : "No root cause" n'est PAS une catégorie de classification.
# C'est un statut technique appliqué automatiquement par le code
# quand cleaning_comment est vide (aucune information disponible
# sur le site) - le modèle n'a jamais à choisir cette valeur.
CATEGORIES = [
    'Coupure ENEO & Baisse de tension',
    'AKTIVCO Coupure ENEO & Baisse de tension',
    'Defaut GE & Power Cabinet',
    'AKTIVCO Defaut GE & Power Cabinet',
    'Sharing', 'Sites strategiques & DataCenter',
    'Sites strategiques, MLL, Pylone, etc.',
    'Projet OCM (HUAWEI)', 'Projet OCM (ZTE, NOKIA, autres projets)',
    'BSS Hardware issue', 'ACCESS-ISSUE', 'MPR issue', 'ODU HS',
    'IP & VLAN', 'fiber AOF', 'fiber CAMTEL',
    'SPARE-ISSUE', 'SPARE-HS', 'SAT', 'EXCLU', 'Warehouse HUAWEI',
]

SYSTEM_PROMPT = """Tu es un expert en analyse d'incidents réseau télécom au NOC
(Network Operations Center). Pour chaque site, tu lis le commentaire nettoyé,
tu identifies le PROBLÈME QUI BLOQUE ACTUELLEMENT la résolution de l'incident,
puis tu classifies ce problème dans exactement une des 21 catégories définies.

PRINCIPE FONDAMENTAL : tu classifies l'ÉTAT ACTUEL du blocage, jamais
l'historique complet de l'incident. Un commentaire est toujours fourni
en entrée - il y a TOUJOURS une information exploitable. Tu dois TOUJOURS
choisir une des 21 catégories ci-dessous : tu n'as jamais la possibilité
de répondre "aucune cause" ou "pas d'information".

Données optionnelles reçues avec le commentaire :
- Owner    : gestionnaire du site
- Topology : configuration électrique du site
- Vendors  : équipementier radio (ZTE, HUAWEI, NOKIA)

===============================================================
ÉTAPE 1 - IDENTIFIER LE PROBLÈME ACTUEL
===============================================================
RÈGLE D'OR : lis les événements du commentaire du PLUS RÉCENT au PLUS ANCIEN.

1. Repère les événements datés dans le commentaire.
2. Parcours-les du plus récent au plus ancien.
3. Pour chaque événement, détermine s'il est BLOQUANT ou RÉSOLU :
   - BLOQUANT : "en attente", "pending", "awaiting", "en cours", "not working",
     "faulty", "defect", "HS", "down" -> C'EST LE PROBLÈME ACTUEL. Arrête-toi ici.
   - RÉSOLU : "delivered", "resolved", "OK", "rétabli", "réparé", "completed",
     "deployed and working" -> Ignore cet événement, continue vers le précédent.
4. Le premier événement BLOQUANT rencontré (en partant du plus récent) est
   LE SEUL problème à classifier. Une cause initiale déjà résolue par un
   événement plus récent n'est JAMAIS reportée, même si elle apparaît dans
   le texte.
5. Si TOUS les événements semblent résolus (aucun événement BLOQUANT
   trouvé), classifie d'après la DERNIÈRE cause technique mentionnée
   dans le commentaire, même résolue - le commentaire reste la seule
   source d'information disponible et doit toujours mener à une des
   21 catégories.

===============================================================
ÉTAPE 2 - CLASSIFIER LE PROBLÈME ACTUEL (ORDRE DE PRIORITÉ STRICT)
===============================================================
Applique ces règles DANS L'ORDRE. Arrête-toi à la première qui matche.

PRIORITÉ 1 - Owner (prime sur tout sauf preuve contraire explicite) :
  Owner contient "MTN" -> Sharing
  Owner contient "OCM" -> Sites strategiques & DataCenter

PRIORITÉ 2 - Projets planifiés (Vendor + mot-clé de projet) :
  Vendor HUAWEI + (swap/migration/commissioning/deployment/upgrade)
    -> Projet OCM (HUAWEI)
  Vendor ZTE/NOKIA + (swap/migration/commissioning/deployment/upgrade)
    -> Projet OCM (ZTE, NOKIA, autres projets)

PRIORITÉ 3 - Équipement technique précis (prime TOUJOURS sur l'énergie,
même si le commentaire mentionne aussi "grid" ou "power" en passant) :
  mpr, eac, mss8, mss, media gateway, mgw, chassis, p8eth, core evo
    -> MPR issue
  odu faulty, odu hs, odu crashed, outdoor unit defect
    -> ODU HS
  frgu, fxdb, frmf, arga, arma, abia, azha, rru/trx/bbu faulty, baseband
    -> BSS Hardware issue
  vlan, ipran, mpls, routing issue, s1 interface
    -> IP & VLAN
  vsat, buc, lnb, satellite, idirect
    -> SAT
  Pour les problèmes de fibre, vérifie CAMTEL avant AOF :
  camtel (ticket/backbone/escalade), même si le commentaire
  mentionne aussi "fiber cut" ou "fo" dans la même phrase
    -> fiber CAMTEL
  sfp, fo cut, fiber cut, fibre coupée, aerial fiber, bbu-rru
  (uniquement si "camtel" n'apparaît PAS dans le commentaire)
    -> fiber AOF

PRIORITÉ 4 - Logistique spare :
  spare hs, spare hors service, spare defect/faulty, DOA (reçu défectueux)
    -> SPARE-HS
  spare non disponible, commande spare, awaiting spare, pas d'ETD/ETA
    -> SPARE-ISSUE
  ETD/ETA défini, hw spms, spms, warehouse huawei, spare en transit
    -> Warehouse HUAWEI

PRIORITÉ 5 - Accès et incidents graves :
  zone rouge, ghost town, bir, accès refusé, bailleur, work permit
    -> ACCESS-ISSUE
  vandalisme, vol, câble coupé/sectionné (acte malveillant), scellés,
  sabotage, mort d'homme, enquête judiciaire
    -> EXCLU

PRIORITÉ 6 - Environnement (sites stratégiques) :
  foudre, pluie forte/violente, intempéries, éboulement, animaux/abeilles,
  vent violent, shelter infiltré
    -> Sites strategiques, MLL, Pylone, etc.

PRIORITÉ 7 - Énergie (UNIQUEMENT si AUCUNE règle ci-dessus n'a matché) :
  Vérifie Owner ET Topology ET mots-clés ENSEMBLE (les trois nécessaires).

  Topology AVEC source alternative (contient GE, Solar, Hybrid, Lithium,
  Gen - ex: Grid-Gen, Hybrid Solar, Bad Grid + GE) :
  Topology SANS source alternative (good grid, good grid no ge, grid only,
  gridonly) :

  Owner CAMUSAT/ESCO + AVEC alternative + (gen/generator/power cabinet/
    rectifier faulty) -> AKTIVCO Defaut GE & Power Cabinet
  Owner CAMUSAT/ESCO + SANS alternative + (grid/eneo/coupure/low voltage/
    power outage) -> AKTIVCO Coupure ENEO & Baisse de tension
  Owner IHS (I_HS, HIS) + AVEC alternative + (gen/generator/power cabinet/
    rectifier faulty) -> Defaut GE & Power Cabinet
  Owner IHS (I_HS, HIS) + SANS alternative + (grid/eneo/coupure/low voltage/
    power outage) -> Coupure ENEO & Baisse de tension

PRIORITÉ 8 - Dernier recours (OBLIGATOIRE - ne jamais laisser sans cause) :
  Si aucune règle ci-dessus ne matche clairement, identifie le signal
  le plus proche dans le commentaire parmi les 21 catégories (même
  un indice faible ou indirect) et choisis la catégorie correspondante.
  Tu dois toujours produire une cause parmi les 21 catégories listées -
  ne réponds jamais autre chose qu'un de ces 21 libellés exacts.

===============================================================
RÈGLES CRITIQUES - RAPPEL
===============================================================
- Lis toujours du plus récent au plus ancien. Le premier bloquant rencontré
  dans cet ordre est LE problème, jamais le premier ni le dernier événement
  du texte brut.
- Une cause initiale résolue n'est jamais reportée si un problème plus
  récent et différent bloque désormais le site.
- Owner prime sur les mots-clés UNIQUEMENT pour Sharing et Sites
  strategiques & DataCenter. Pour l'énergie, Owner + Topology + mots-clés
  sont TOUS nécessaires ensemble.
- Un équipement technique précis (MPR, ODU, BSS, fibre, IP, SAT) prime
  toujours sur une cause énergie générique.
- Pour la fibre, CAMTEL est toujours vérifié avant AOF.
- Un mot-clé court doit matcher un mot entier, jamais une sous-chaîne
  (ex: "fo" ne doit jamais matcher dans "information").
- Tu dois TOUJOURS choisir une des 21 catégories listées. Il n'existe
  aucune réponse de type "pas de cause" ou "information insuffisante".

===============================================================
EXEMPLE
===============================================================
Commentaire : "10/06: Grid outage - 12/06: Gen deployed, site stable -
15/06: Spare HS, waiting replacement"
Lecture du plus récent : 15/06 "Spare HS, waiting replacement" = BLOQUANT
-> STOP ici. Résultat : SPARE-HS

===============================================================
LISTE DES 21 CATÉGORIES EXACTES
===============================================================
1. Coupure ENEO & Baisse de tension
2. AKTIVCO Coupure ENEO & Baisse de tension
3. Defaut GE & Power Cabinet
4. AKTIVCO Defaut GE & Power Cabinet
5. Sharing
6. Sites strategiques & DataCenter
7. Sites strategiques, MLL, Pylone, etc.
8. Projet OCM (HUAWEI)
9. Projet OCM (ZTE, NOKIA, autres projets)
10. BSS Hardware issue
11. ACCESS-ISSUE
12. MPR issue
13. ODU HS
14. IP & VLAN
15. fiber AOF
16. fiber CAMTEL
17. SPARE-ISSUE
18. SPARE-HS
19. SAT
20. EXCLU
21. Warehouse HUAWEI

FORMAT DE RÉPONSE - TOUJOURS CE FORMAT EXACT :
Analyse : [type de problème identifié + équipements concernés + statut en cours ou clôturé]
Cause : [libellé exact parmi les 21 catégories ci-dessus]"""

# -- CHARGEMENT DU MODÈLE - UNE SEULE FOIS AU DÉMARRAGE --------
print("[...] Chargement du tokenizer...")
QWEN_CHAT_TEMPLATE = (
    "{% for message in messages %}"
    "{% if message['role'] == 'system' %}<|im_start|>system\n{{ message['content'] }}<|im_end|>\n"
    "{% elif message['role'] == 'user' %}<|im_start|>user\n{{ message['content'] }}<|im_end|>\n"
    "{% elif message['role'] == 'assistant' %}<|im_start|>assistant\n{{ message['content'] }}<|im_end|>\n"
    "{% endif %}{% endfor %}"
    "{% if add_generation_prompt %}<|im_start|>assistant\n{% endif %}"
)

tokenizer = AutoTokenizer.from_pretrained(
    MODELE_PATH, trust_remote_code=True, local_files_only=True
)
tokenizer.pad_token = tokenizer.eos_token
tokenizer.padding_side = "right"
tokenizer.chat_template = QWEN_CHAT_TEMPLATE
print("[OK] Tokenizer chargé")

print("[...] Chargement du modèle (GPU)...")
quant_config = BitsAndBytesConfig(
    load_in_4bit=True,
    bnb_4bit_compute_dtype=torch.float16,
    bnb_4bit_quant_type="nf4",
    bnb_4bit_use_double_quant=True,
)
model = AutoModelForCausalLM.from_pretrained(
    MODELE_PATH,
    quantization_config=quant_config,
    device_map="auto",
    trust_remote_code=True,
    local_files_only=True,
)
model.eval()
print(f"[OK] Modèle prêt - VRAM utilisée : {torch.cuda.memory_allocated()/1024**3:.2f} GB")


# -- FONCTIONS DE CLASSIFICATION -------------------------------
def propre(v):
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s.lower() in ["nan", "none", "null", "", "-", "n/a"] else s


def construire_prompt(commentaire, owner="", topology="", vendors=""):
    user = f'Commentaire : "{commentaire.strip()}"'
    if propre(owner):
        user += f"\nOwner : {propre(owner)}"
    if propre(topology):
        user += f"\nTopology : {propre(topology)}"
    if propre(vendors):
        user += f"\nVendors : {propre(vendors)}"
    return user


def normaliser_cause(cause):
    cause = cause.strip()
    for prefix in ["cause :", "cause:"]:
        if cause.lower().startswith(prefix):
            cause = cause[len(prefix):].strip()
    for cat in CATEGORIES:
        if cat.lower() == cause.lower():
            return cat
    for cat in CATEGORIES:
        if cat.lower() in cause.lower():
            return cat
    return cause


def predire(commentaire, owner="", topology="", vendors=""):
    user_content = construire_prompt(commentaire, owner, topology, vendors)
    messages = [
        {"role": "system", "content": SYSTEM_PROMPT},
        {"role": "user", "content": user_content},
    ]
    prompt = tokenizer.apply_chat_template(
        messages, tokenize=False, add_generation_prompt=True
    )
    inputs = tokenizer(prompt, return_tensors="pt", truncation=True, max_length=3500)
    inputs = {k: v.to(model.device) for k, v in inputs.items()}

    with torch.no_grad():
        outputs = model.generate(
            **inputs,
            max_new_tokens=150,
            do_sample=False,
            repetition_penalty=1.05,
            pad_token_id=tokenizer.eos_token_id,
            eos_token_id=tokenizer.eos_token_id,
        )

    reponse = tokenizer.decode(
        outputs[0][inputs["input_ids"].shape[1]:], skip_special_tokens=True
    ).strip()

    return _post_traiter_reponse(reponse, owner)


# -- BATCH SIZE - ajustable selon la VRAM disponible du GPU loué --
# 24 GB VRAM (RTX 3090/4090/A5000) : 8 est une valeur sûre.
# Si OOM (Out Of Memory) au démarrage du traitement, réduire à 4.
# Si beaucoup de VRAM libre (ex: A6000 48GB), monter à 12-16 pour
# encore accélérer.
BATCH_SIZE = 8


def _post_traiter_reponse(reponse, owner=""):
    def extraire(texte, debut, fin=None):
        if debut not in texte:
            return ""
        c = texte.split(debut, 1)[1]
        if fin and fin in c:
            c = c.split(fin, 1)[0]
        return c.strip()

    analyse = extraire(reponse, "Analyse :", "Cause :")
    cause_brute = extraire(reponse, "Cause :").split("\n")[0].strip()
    cause = normaliser_cause(cause_brute)

    if not cause or cause not in CATEGORIES:
        for cat in CATEGORIES:
            if cat.lower() in reponse.lower():
                cause = cat
                break

    owner_lower = propre(owner).lower()
    if owner_lower in ["camusat", "esco"]:
        if cause == "Coupure ENEO & Baisse de tension":
            cause = "AKTIVCO Coupure ENEO & Baisse de tension"
        elif cause == "Defaut GE & Power Cabinet":
            cause = "AKTIVCO Defaut GE & Power Cabinet"

    return {"analyse": analyse, "cause": cause, "reponse_complete": reponse}


def predire_batch(lignes):
    """
    Traite plusieurs commentaires en un seul passage GPU.
    lignes : liste de dicts {commentaire, owner, topology, vendors}
    Retourne : liste de résultats dans le même ordre.
    """
    if not lignes:
        return []

    prompts = []
    for ligne in lignes:
        user_content = construire_prompt(
            ligne["commentaire"], ligne.get("owner", ""),
            ligne.get("topology", ""), ligne.get("vendors", "")
        )
        messages = [
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user", "content": user_content},
        ]
        prompt = tokenizer.apply_chat_template(
            messages, tokenize=False, add_generation_prompt=True
        )
        prompts.append(prompt)

    # Tokenisation par batch avec padding à gauche (requis pour
    # une génération causale correcte avec un batch de tailles variées)
    tokenizer.padding_side = "left"
    inputs = tokenizer(
        prompts, return_tensors="pt", truncation=True,
        max_length=3500, padding=True,
    )
    inputs = {k: v.to(model.device) for k, v in inputs.items()}

    with torch.no_grad():
        outputs = model.generate(
            **inputs,
            max_new_tokens=150,
            do_sample=False,
            repetition_penalty=1.05,
            pad_token_id=tokenizer.eos_token_id,
            eos_token_id=tokenizer.eos_token_id,
        )
    tokenizer.padding_side = "right"  # restaurer pour /chat et /predire

    resultats = []
    longueur_entree = inputs["input_ids"].shape[1]
    for i, ligne in enumerate(lignes):
        reponse = tokenizer.decode(
            outputs[i][longueur_entree:], skip_special_tokens=True
        ).strip()
        resultats.append(_post_traiter_reponse(reponse, ligne.get("owner", "")))

    return resultats


# -- Détection des sites impactés ------------------------------
# Un site est "impacté" quand son indisponibilité résulte d'un
# problème sur un AUTRE site (site parent), identifié par l'une
# des formulations suivantes dans le commentaire. Le code du site
# parent suit directement le préfixe et fait toujours 7 caractères
# au total (4 caractères de préfixe + 3 caractères, ex: CTR_852),
# reproduisant la logique de la formule Excel
# =MID(N2, SEARCH("CTR_", N2), 7) utilisée en amont sur ces fichiers.
PREFIXES_SITES = ["CTR_", "SUO_", "SUD_", "NRO_", "ADM_", "NRD_", "EXN_", "LIT_", "EST_", "OST_"]
PATTERN_CODE_SITE = re.compile(
    r"(?:" + "|".join(re.escape(p) for p in PREFIXES_SITES) + r")[A-Z0-9]{3}",
    re.IGNORECASE,
)

MOTS_CLES_IMPACT = [
    r"indisponibilit[eé] des services? voix\s*[&et]+\s*data dans la localit[eé] de",
    r"impacted?\s+by",
    r"impact[eé]r?\s+par",
    r"incident\s+at",
    r"probable\s+probl[eè]me\s+de.*?sur\s+le\s+site\s+de",
    r"affect(?:ed)?\s+by(?:\s+power\s+fluctuations\s+at)?",
]
PATTERN_IMPACT = re.compile("|".join(MOTS_CLES_IMPACT), re.IGNORECASE)


def est_impacte(com):
    """
    Un site n'est considéré comme réellement impacté par un AUTRE
    site QUE SI un mot-clé déclencheur ET un code site structuré
    sont tous les deux présents dans le commentaire. Le mot-clé
    seul ne suffit pas : "Site impacté par de fortes intempéries"
    contient "impacté par" mais ne désigne aucun site parent - ce
    n'est pas un cas d'impact, c'est une cause environnementale
    normale à classifier directement par le modèle.
    """
    if not PATTERN_IMPACT.search(com):
        return False
    return bool(PATTERN_CODE_SITE.search(com))


def extraire_site_impactant(com):
    """
    Cherche le code du site parent dans le commentaire : préfixe
    (CTR_, SUO_, etc.) suivi de 3 caractères, soit 7 caractères au
    total - identique à la formule Excel MID(N2, SEARCH(prefix, N2), 7)
    utilisée en amont. Recherche sur tout le commentaire plutôt que
    juste après le mot-clé déclencheur, pour rester robuste face aux
    variations de formulation.
    Retourne (code_site, statut) où statut est CODE_TROUVE ou PAS_DE_CODE.
    """
    if not com:
        return "", "PAS_DE_CODE"
    match = PATTERN_CODE_SITE.search(com)
    if match:
        return match.group(0).strip().upper(), "CODE_TROUVE"
    return "", "PAS_DE_CODE"


def trouver_site_parent(code, df, col_site, col_nom_physique):
    """
    Recherche le code du site parent dans les colonnes Codesite
    ET Nom du Physique (le code peut apparaître dans l'une ou
    l'autre selon le fichier).
    Retourne l'index de la ligne trouvée, ou None.
    """
    if not code:
        return None

    code_norm = code.strip().upper()

    masque_site = df[col_site].astype(str).str.strip().str.upper() == code_norm
    if masque_site.any():
        return df[masque_site].index[0]

    if col_nom_physique:
        masque_nom = df[col_nom_physique].astype(str).str.upper().str.contains(
            re.escape(code_norm), na=False
        )
        if masque_nom.any():
            return df[masque_nom].index[0]

    return None


# ==============================================================
# ROUTE 1 - /classifier : traiter un fichier Excel complet
# ==============================================================
import threading
import uuid


# -- SYSTEME DE TACHES EN ARRIERE-PLAN --------------------------
# Le proxy RunPod (Cloudflare) coupe automatiquement toute requête
# HTTP qui reste ouverte plus d'environ 100 secondes sans réponse.
# Comme le traitement d'un fichier complet dure plusieurs minutes,
# on ne peut plus faire attendre la requête HTTP jusqu'à la fin :
# le serveur démarre le traitement en arrière-plan et répond
# IMMÉDIATEMENT avec un identifiant de tâche. Le client interroge
# ensuite régulièrement /classifier/statut/{task_id} (réponse quasi
# instantanée, jamais bloquée par le proxy) jusqu'à ce que le
# traitement soit terminé, puis télécharge le résultat séparément.
#
# Le fichier n'est jamais découpé : il est toujours traité en une
# seule fois, par le même code de classification qu'auparavant.
TACHES = {}  # task_id -> dict d'état de la tâche
TACHES_LOCK = threading.Lock()


def _maj_tache(task_id, **kwargs):
    with TACHES_LOCK:
        if task_id in TACHES:
            TACHES[task_id].update(kwargs)


def traiter_fichier_arriere_plan(task_id, tmp_path):
    """
    Reprend exactement la même logique de traitement qu'avant
    (détection colonnes, classification par batch, sites impactés,
    écriture openpyxl) mais s'exécute dans un thread séparé, en
    mettant à jour l'état de la tâche au fur et à mesure pour que
    /classifier/statut/{task_id} puisse renseigner la progression.
    """
    t_debut = time.time()
    NOM_FEUILLE = "daily break down"

    try:
        df = pd.read_excel(tmp_path, sheet_name=NOM_FEUILLE, dtype=str)
    except Exception as e:
        _maj_tache(task_id, statut="erreur",
                   message=f"Impossible de lire la feuille '{NOM_FEUILLE}' : {e}")
        os.unlink(tmp_path)
        return

    col_verif = next((c for c in df.columns if "cleaning_comment" in c.lower()), None)
    if col_verif is None:
        col_verif = next((c for c in df.columns if "verif" in c.lower()), None)
    col_site = next((c for c in df.columns if "codesite" in c.lower()), None)
    col_nom_physique = next((c for c in df.columns if "nom du physique" in c.lower()), None)
    col_own = next((c for c in df.columns if c.lower() == "owner"), None)
    col_top = next((c for c in df.columns if "topology" in c.lower()), None)
    col_ven = next((c for c in df.columns if "vendor" in c.lower()), None)
    col_cause = next((c for c in df.columns if c.lower() == "cause"), None)
    col_impacts = next((c for c in df.columns if "impact" in c.lower()), None)

    manquantes = []
    if col_verif is None: manquantes.append("cleaning_comment")
    if col_site is None: manquantes.append("codesite")
    if col_cause is None: manquantes.append("CAUSE")
    if col_impacts is None: manquantes.append("impacts")

    if manquantes:
        _maj_tache(
            task_id, statut="erreur",
            message=f"Colonnes requises introuvables : {', '.join(manquantes)}. "
                    f"Colonnes disponibles : {list(df.columns)}",
        )
        os.unlink(tmp_path)
        return

    idx_impactes = []
    idx_a_traiter = []
    resultat_cause = {}
    resultat_impact = {}

    for idx, row in df.iterrows():
        com = propre(str(row.get(col_verif, "")))
        if not com:
            resultat_cause[idx] = "No root cause"
            continue
        if est_impacte(com):
            idx_impactes.append(idx)
            continue
        idx_a_traiter.append(idx)

    total_lignes_a_traiter = len(idx_a_traiter)
    _maj_tache(
        task_id, statut="en_cours",
        total=total_lignes_a_traiter, traitees=0,
        message=f"{total_lignes_a_traiter} lignes à classifier, "
                f"{len(idx_impactes)} cas impactés",
    )
    print(f"[BATCH] {total_lignes_a_traiter} lignes à classifier, "
          f"{len(idx_impactes)} cas impactés, batch_size={BATCH_SIZE}")

    for debut in range(0, len(idx_a_traiter), BATCH_SIZE):
        lot_idx = idx_a_traiter[debut: debut + BATCH_SIZE]
        lot_lignes = []
        for idx in lot_idx:
            row = df.loc[idx]
            lot_lignes.append({
                "commentaire": propre(str(row.get(col_verif, ""))),
                "owner": propre(str(row.get(col_own, ""))) if col_own else "",
                "topology": propre(str(row.get(col_top, ""))) if col_top else "",
                "vendors": propre(str(row.get(col_ven, ""))) if col_ven else "",
            })

        try:
            resultats_lot = predire_batch(lot_lignes)
        except Exception as e:
            print(f"[ATTENTION] Erreur sur le batch démarrant à {debut} : {e}")
            resultats_lot = [{"cause": "ERREUR", "analyse": ""} for _ in lot_idx]

        for idx, res in zip(lot_idx, resultats_lot):
            resultat_cause[idx] = res["cause"]

        traites = min(debut + BATCH_SIZE, total_lignes_a_traiter)
        ecoule = time.time() - t_debut
        _maj_tache(task_id, traitees=traites)
        print(f"   {traites}/{total_lignes_a_traiter} lignes | "
              f"{ecoule:.0f}s écoulées | "
              f"{ecoule/max(traites,1):.2f}s/ligne moyen")

    code_parent_par_idx = {}
    for idx in idx_impactes:
        com = propre(str(df.loc[idx].get(col_verif, "")))
        code, _ = extraire_site_impactant(com)
        code_parent_par_idx[idx] = code

    idx_impactes_restants = list(idx_impactes)
    nb_parent_trouve = 0

    for _ in range(len(idx_impactes) + 1):
        if not idx_impactes_restants:
            break
        encore_en_attente = []
        progresse = False

        for idx in idx_impactes_restants:
            code = code_parent_par_idx.get(idx, "")
            idx_parent = trouver_site_parent(code, df, col_site, col_nom_physique) if code else None

            if idx_parent is None:
                resultat_cause[idx] = "cause site parent pas retrouver"
                resultat_impact[idx] = f"Site impacté par {code or 'site inconnu'} - parent introuvable dans le fichier"
                progresse = True
            elif idx_parent in resultat_cause:
                resultat_cause[idx] = resultat_cause[idx_parent]
                resultat_impact[idx] = f"Site impacté par {code}"
                nb_parent_trouve += 1
                progresse = True
            else:
                encore_en_attente.append(idx)

        idx_impactes_restants = encore_en_attente
        if not progresse:
            break

    for idx in idx_impactes_restants:
        code = code_parent_par_idx.get(idx, "")
        resultat_cause[idx] = "cause site parent pas retrouver"
        resultat_impact[idx] = f"Site impacté par {code or 'site inconnu'} - parent introuvable dans le fichier"

    nb_parent_introuvable = len(idx_impactes) - nb_parent_trouve
    print(f"[LIEN] Sites impactés : {nb_parent_trouve} parent(s) trouvé(s), "
          f"{nb_parent_introuvable} parent(s) introuvable(s)")

    try:
        wb = load_workbook(tmp_path)
        ws = wb[NOM_FEUILLE]

        en_tetes = {}
        for cell in ws[1]:
            if cell.value is not None:
                en_tetes[str(cell.value).strip().lower()] = cell.column

        col_idx_cause = en_tetes.get("cause")
        col_idx_impacts = next((v for k, v in en_tetes.items() if "impact" in k), None)

        if col_idx_cause is None or col_idx_impacts is None:
            _maj_tache(
                task_id, statut="erreur",
                message="Colonnes CAUSE/impacts repérées par pandas mais "
                        "introuvables dans le fichier openpyxl.",
            )
            os.unlink(tmp_path)
            return

        for idx in df.index:
            ligne_excel = idx + 2
            cause = resultat_cause.get(idx, "No root cause")
            ws.cell(row=ligne_excel, column=col_idx_cause, value=cause)
            if idx in resultat_impact:
                ws.cell(row=ligne_excel, column=col_idx_impacts, value=resultat_impact[idx])

        result_path = tmp_path.replace(".xlsx", "_resultat.xlsx")
        wb.save(result_path)
        os.unlink(tmp_path)

        duree = time.time() - t_debut
        print(f"[OK] Fichier traité en {duree/60:.1f} min ({len(df)} lignes, "
              f"{total_lignes_a_traiter} classifiées par le modèle)")

        _maj_tache(
            task_id, statut="termine", traitees=total_lignes_a_traiter,
            resultat_path=result_path,
            message=f"Traitement terminé en {duree/60:.1f} min",
        )

    except Exception as e:
        _maj_tache(task_id, statut="erreur", message=f"Erreur lors de l'écriture du fichier : {e}")
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)


@app.post("/classifier/demarrer")
async def classifier_demarrer(file: UploadFile = File(...)):
    """
    Reçoit le fichier complet (en une seule fois, sans découpage)
    et lance immédiatement le traitement dans un thread séparé.
    Répond en moins d'une seconde avec un identifiant de tâche,
    bien avant que le proxy ne puisse couper la connexion.
    """
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Le fichier doit être un .xlsx ou .xls")

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        contenu = await file.read()
        tmp.write(contenu)
        tmp_path = tmp.name

    task_id = str(uuid.uuid4())
    with TACHES_LOCK:
        TACHES[task_id] = {
            "statut": "en_attente",
            "total": 0,
            "traitees": 0,
            "message": "Tâche reçue, démarrage du traitement...",
            "resultat_path": None,
        }

    thread = threading.Thread(
        target=traiter_fichier_arriere_plan,
        args=(task_id, tmp_path),
        daemon=True,
    )
    thread.start()

    return {"task_id": task_id}


@app.get("/classifier/statut/{task_id}")
def classifier_statut(task_id: str):
    """
    Réponse quasi instantanée donnant l'état d'avancement.
    Le client (Streamlit) interroge cette route toutes les
    quelques secondes pour suivre la progression sans jamais
    laisser une requête HTTP ouverte trop longtemps.
    """
    with TACHES_LOCK:
        tache = TACHES.get(task_id)

    if tache is None:
        raise HTTPException(404, "Tâche introuvable.")

    pourcentage = 0
    if tache["total"] > 0:
        pourcentage = round(tache["traitees"] / tache["total"] * 100, 1)

    return {
        "statut": tache["statut"],
        "total": tache["total"],
        "traitees": tache["traitees"],
        "pourcentage": pourcentage,
        "message": tache["message"],
    }


@app.get("/classifier/resultat/{task_id}")
def classifier_resultat(task_id: str):
    """
    Renvoie le fichier final, uniquement disponible une fois le
    statut passé à "termine".
    """
    with TACHES_LOCK:
        tache = TACHES.get(task_id)

    if tache is None:
        raise HTTPException(404, "Tâche introuvable.")

    if tache["statut"] == "erreur":
        raise HTTPException(500, tache["message"])

    if tache["statut"] != "termine":
        raise HTTPException(409, "Le traitement n'est pas encore terminé.")

    result_path = tache["resultat_path"]
    if not result_path or not os.path.exists(result_path):
        raise HTTPException(500, "Fichier résultat introuvable.")

    return FileResponse(
        result_path,
        filename="breakdown_classifie.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ==============================================================
# ROUTE 2 - /chat : dialogue libre avec le modèle
# ==============================================================
class ChatMessage(BaseModel):
    role: str
    contenu: str


class ChatRequest(BaseModel):
    commentaire: str
    owner: Optional[str] = ""
    topology: Optional[str] = ""
    vendors: Optional[str] = ""


@app.post("/chat")
def chat_route(data: ChatRequest):
    if not data.commentaire or not data.commentaire.strip():
        raise HTTPException(400, "Le commentaire ne peut pas être vide.")

    res = predire(data.commentaire, data.owner, data.topology, data.vendors)
    return {
        "cause": res["cause"],
        "analyse": res["analyse"],
    }


# ==============================================================
# ROUTE 3 - /health : vérification de l'état du serveur
# ==============================================================
@app.get("/health")
def health():
    return {
        "status": "ok",
        "modele": "Qwen2.5-3B fine-tuné",
        "gpu": torch.cuda.get_device_name(0) if torch.cuda.is_available() else "aucun",
        "vram_utilisee_gb": round(torch.cuda.memory_allocated() / 1024**3, 2),
    }


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)